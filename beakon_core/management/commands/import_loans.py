"""Import the workbook tab `07 Loan Master` into the Loan master table.

Usage:
    python manage.py import_loans [--workbook PATH] [--organization-id ID] [--dry-run]

Idempotent: re-running updates existing rows in place keyed on
``(organization, loan_id)``. Resolves four ``Default_*_Account`` codes to
``Account`` foreign keys when the codes match existing accounts.

Three columns (``Spread_Bps``, ``Scheduled_Principal_Amount``,
``Default_FX_Gain_Loss_Account``) are stored as text because the workbook
has them auto-coerced into Excel dates. They round-trip raw so Thomas can
fix the cells without touching code.
"""
from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from beakon_core.models import Account, Loan
from organizations.models import Organization


DEFAULT_WORKBOOK = r"D:\Thomas\2026 04 17-DRAFT-CoA-Wealth management v2.xlsx"
SHEET_NAME = "07 Loan Master"

DATE_FIELDS = {
    "start_date", "first_accrual_date", "maturity_date",
    "next_reset_date", "next_interest_payment_date", "next_principal_payment_date",
    "effective_from", "effective_to",
}
DECIMAL_FIELDS = {
    "principal_original", "current_principal_outstanding", "fixed_rate",
}
BOOL_FIELDS = {
    "related_party_flag", "bullet_flag", "prepayment_allowed_flag",
    "capitalized_interest_flag", "accrual_required_flag", "fx_remeasure_flag",
    "approval_required_flag", "manual_override_allowed_flag",
    "source_document_required_flag",
}
# Workbook columns that Excel coerced to dates — store raw text until cleaned.
RAW_TEXT_FIELDS = {
    "spread_bps", "scheduled_principal_amount", "default_fx_gain_loss_account_code",
}

COLUMN_TO_FIELD: dict[str, str] = {
    "Loan_ID": "loan_id",
    "Loan_Name": "loan_name",
    "Loan_Type": "loan_type",
    "Loan_Side": "loan_side",
    "Status": "status",
    "Borrower_or_Lender_Code": "borrower_or_lender_code",
    "Related_Party_Flag": "related_party_flag",
    "Facility_Reference": "facility_reference",
    "Internal_Reference": "internal_reference",
    "Loan_Currency": "loan_currency",
    "Principal_Original": "principal_original",
    "Current_Principal_Outstanding": "current_principal_outstanding",
    "Interest_Rate_Type": "interest_rate_type",
    "Fixed_Rate": "fixed_rate",
    "Reference_Rate_Code": "reference_rate_code",
    "Spread_Bps": "spread_bps",
    "Interest_Reset_Frequency": "interest_reset_frequency",
    "Interest_Payment_Frequency": "interest_payment_frequency",
    "Day_Count_Convention": "day_count_convention",
    "Start_Date": "start_date",
    "First_Accrual_Date": "first_accrual_date",
    "Maturity_Date": "maturity_date",
    "Next_Reset_Date": "next_reset_date",
    "Next_Interest_Payment_Date": "next_interest_payment_date",
    "Next_Principal_Payment_Date": "next_principal_payment_date",
    "Repayment_Type": "repayment_type",
    "Amortization_Method": "amortization_method",
    "Bullet_Flag": "bullet_flag",
    "Scheduled_Principal_Amount": "scheduled_principal_amount",
    "Prepayment_Allowed_Flag": "prepayment_allowed_flag",
    "Capitalized_Interest_Flag": "capitalized_interest_flag",
    "Reporting_Portfolio_Code": "reporting_portfolio_code",
    "Collateral_Link_Type": "collateral_link_type",
    "Collateral_Link_ID": "collateral_link_id",
    "Current_Noncurrent_Split_Method": "current_noncurrent_split_method",
    "Valuation_Basis": "valuation_basis",
    "Impairment_Method": "impairment_method",
    "Accrual_Required_Flag": "accrual_required_flag",
    "FX_Remeasure_Flag": "fx_remeasure_flag",
    "Default_Principal_Account": "default_principal_account_code",
    "Default_Interest_Income_Account": "default_interest_income_account_code",
    "Default_Interest_Expense_Account": "default_interest_expense_account_code",
    "Default_FX_Gain_Loss_Account": "default_fx_gain_loss_account_code",
    "Approval_Required_Flag": "approval_required_flag",
    "Manual_Override_Allowed_Flag": "manual_override_allowed_flag",
    "Source_Document_Required_Flag": "source_document_required_flag",
    "Effective_From": "effective_from",
    "Effective_To": "effective_to",
    "Notes": "notes",
}

ACCOUNT_FK_FIELDS = [
    ("default_principal_account_code", "default_principal_account"),
    ("default_interest_income_account_code", "default_interest_income_account"),
    ("default_interest_expense_account_code", "default_interest_expense_account"),
    ("default_fx_gain_loss_account_code", "default_fx_gain_loss_account"),
]


class Command(BaseCommand):
    help = "Import tab 07 Loan Master from Thomas's wealth-management workbook."

    def add_arguments(self, parser):
        parser.add_argument("--workbook", default=DEFAULT_WORKBOOK)
        parser.add_argument("--organization-id", type=int)
        parser.add_argument("--dry-run", action="store_true")

    def handle(self, *args, **options):
        path = Path(options["workbook"])
        if not path.exists():
            raise CommandError(f"Workbook not found: {path}")

        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise CommandError("openpyxl not installed.") from exc

        org = _resolve_organization(options.get("organization_id"))
        wb = load_workbook(path, data_only=True, read_only=False)
        if SHEET_NAME not in wb.sheetnames:
            raise CommandError(f"Sheet '{SHEET_NAME}' not found in workbook.")

        ws = wb[SHEET_NAME]
        rows = _rows_by_header(ws)

        account_lookup = {a.code: a for a in Account.objects.filter(organization=org)}

        stats = {
            "created": 0, "updated": 0,
            "skipped_no_id": 0, "skipped_explanation_row": 0,
            "linked_principal": 0, "linked_interest_income": 0,
            "linked_interest_expense": 0, "linked_fx_gain_loss": 0,
        }
        errors: list[str] = []

        with transaction.atomic():
            for row in rows:
                loan_id = _text(row.get("Loan_ID"))
                if not loan_id:
                    stats["skipped_no_id"] += 1
                    continue
                if not loan_id.startswith("LOAN_"):
                    stats["skipped_explanation_row"] += 1
                    continue

                defaults = self._build_defaults(row, account_lookup, errors, stats)
                _, created = Loan.objects.update_or_create(
                    organization=org,
                    loan_id=loan_id,
                    defaults=defaults,
                )
                stats["created" if created else "updated"] += 1

            if options["dry_run"]:
                transaction.set_rollback(True)

        prefix = "DRY RUN parsed" if options["dry_run"] else "Imported"
        self.stdout.write(self.style.SUCCESS(
            f"{prefix} Loan rows for org {org.id} ({org.name})"
        ))
        for k, v in stats.items():
            self.stdout.write(f"  {k}: {v}")
        if errors:
            self.stdout.write(self.style.WARNING(f"{len(errors)} warning(s):"))
            for e in errors[:20]:
                self.stdout.write(f"  - {e}")
            if len(errors) > 20:
                self.stdout.write(f"  ... and {len(errors) - 20} more")

    # ─────────────── per-row mapping ─────────────── #

    def _build_defaults(self, row: dict[str, Any], account_lookup, errors, stats) -> dict:
        defaults: dict[str, Any] = {}
        metadata: dict[str, Any] = {}

        for column, value in row.items():
            if column == "_row":
                continue
            if column in COLUMN_TO_FIELD:
                field = COLUMN_TO_FIELD[column]
                defaults[field] = self._coerce(field, value)
            else:
                if value not in (None, ""):
                    metadata[column] = _json_safe(value)

        # Normalise enum-ish strings.
        for key in ("loan_type", "loan_side", "status", "interest_rate_type",
                    "repayment_type", "valuation_basis", "collateral_link_type"):
            if defaults.get(key):
                defaults[key] = defaults[key].upper()

        # Resolve default-account FK columns from the workbook codes. We only
        # link when the code is a clean account number (skip the corrupted-
        # date column for the FX-gain-loss account; the raw text is preserved
        # in the *_code field for later cleanup).
        for code_field, fk_field in ACCOUNT_FK_FIELDS:
            code = _text(defaults.get(code_field))
            if not code or not code.replace(".", "").isdigit():
                continue
            account = account_lookup.get(code)
            if account is not None:
                defaults[fk_field] = account
                # Track linked-vs-unlinked counts in stats
                short = fk_field.replace("default_", "").replace("_account", "")
                key = f"linked_{short}"
                if key in stats:
                    stats[key] += 1
            else:
                errors.append(
                    f"Loan '{defaults.get('loan_id')}' references "
                    f"unknown {code_field} '{code}'."
                )

        defaults["workbook_metadata"] = metadata
        return defaults

    def _coerce(self, field: str, value: Any) -> Any:
        if field in RAW_TEXT_FIELDS:
            return _raw_text(value)
        if field in DATE_FIELDS:
            return _date(value)
        if field in DECIMAL_FIELDS:
            return _decimal(value, default=None)
        if field in BOOL_FIELDS:
            return _bool(value, default=False)
        return _text(value)


# ────────────────────────── helpers ────────────────────────── #


def _resolve_organization(organization_id: int | None) -> Organization:
    if organization_id:
        try:
            return Organization.objects.get(id=organization_id)
        except Organization.DoesNotExist as exc:
            raise CommandError(f"Organization id {organization_id} does not exist.") from exc

    orgs = list(Organization.objects.all()[:2])
    if not orgs:
        raise CommandError("No Organization exists. Pass --organization-id.")
    if len(orgs) > 1:
        raise CommandError("Multiple organizations exist. Re-run with --organization-id.")
    return orgs[0]


def _rows_by_header(ws) -> list[dict[str, Any]]:
    headers = [_text(ws.cell(1, col).value) for col in range(1, ws.max_column + 1)]
    rows: list[dict[str, Any]] = []
    for row_no in range(2, ws.max_row + 1):
        row: dict[str, Any] = {"_row": row_no}
        has_value = False
        for idx, header in enumerate(headers, start=1):
            value = _value(ws.cell(row_no, idx).value)
            if value not in ("", None):
                has_value = True
            if header:
                row[header] = value
        if has_value:
            rows.append(row)
    return rows


def _value(value: Any) -> Any:
    if isinstance(value, str):
        return value.strip()
    return value


def _text(value: Any) -> str:
    value = _value(value)
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, Decimal):
        return str(value)
    return str(value).strip()


def _raw_text(value: Any) -> str:
    """Preserve a workbook value as text — including Excel-corrupted dates."""
    value = _value(value)
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return _text(value)


def _bool(value: Any, *, default: bool) -> bool:
    text = _text(value).lower()
    if not text:
        return default
    if text in {"yes", "y", "true", "1", "active"}:
        return True
    if text in {"no", "n", "false", "0", "inactive"}:
        return False
    return default


def _decimal(value: Any, *, default):
    if value in ("", None):
        return default
    try:
        return Decimal(str(value))
    except Exception:
        return default


def _date(value: Any):
    if value in ("", None):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    try:
        return datetime.strptime(_text(value), "%Y-%m-%d").date()
    except ValueError:
        return None


def _json_safe(value: Any) -> Any:
    if isinstance(value, dict):
        return {str(k): _json_safe(v) for k, v in value.items() if not str(k).startswith("_")}
    if isinstance(value, list):
        return [_json_safe(v) for v in value]
    if isinstance(value, tuple):
        return [_json_safe(v) for v in value]
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, Decimal):
        return str(value)
    return value
