"""Import the workbook tab `14_Portfolio_Master` into the Portfolio table.

Usage:
    python manage.py import_portfolios [--workbook PATH] [--organization-id ID] [--dry-run]

Idempotent: re-running updates existing rows in place keyed on
``(organization, portfolio_id)``. Two-pass import:
  1. upsert all rows (with ``parent_portfolio_workbook_id`` populated but
     ``parent`` left null)
  2. resolve ``parent`` FK from the workbook codes
"""
from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from beakon_core.models import Portfolio
from organizations.models import Organization


DEFAULT_WORKBOOK = r"D:\Thomas\2026 04 17-DRAFT-CoA-Wealth management v2.xlsx"
SHEET_NAME = "14_Portfolio_Master"

DATE_FIELDS = {"open_date", "close_date"}
BOOL_FIELDS = {
    "discretionary_flag", "consolidation_flag", "net_worth_inclusion_flag",
    "performance_report_flag", "posting_allowed_flag", "active_flag",
    "approval_required_flag", "source_document_required_flag",
}

COLUMN_TO_FIELD: dict[str, str] = {
    "Portfolio_ID": "portfolio_id",
    "Portfolio_Name": "portfolio_name",
    "Short_Name": "short_name",
    "Portfolio_Type": "portfolio_type",
    "Portfolio_Subtype": "portfolio_subtype",
    "Owner_Type": "owner_type",
    "Owner_ID": "owner_id",
    "Primary_Related_Party_ID": "primary_related_party_id",
    "Linked_Custodian_ID": "linked_custodian_id",
    "Base_Currency": "base_currency",
    "Reporting_Currency": "reporting_currency",
    "Country_Code": "country_code",
    "Jurisdiction_Code": "jurisdiction_code",
    "Strategy_Code": "strategy_code",
    "Asset_Allocation_Profile": "asset_allocation_profile",
    "Discretionary_Flag": "discretionary_flag",
    "Consolidation_Flag": "consolidation_flag",
    "Net_Worth_Inclusion_Flag": "net_worth_inclusion_flag",
    "Performance_Report_Flag": "performance_report_flag",
    "Posting_Allowed_Flag": "posting_allowed_flag",
    "Status": "status",
    "Active_Flag": "active_flag",
    "Open_Date": "open_date",
    "Close_Date": "close_date",
    "Parent_Portfolio_ID": "parent_portfolio_workbook_id",
    "Reporting_Group": "reporting_group",
    "Approval_Required_Flag": "approval_required_flag",
    "Source_Document_Required_Flag": "source_document_required_flag",
    "Notes": "notes",
}


class Command(BaseCommand):
    help = "Import tab 14_Portfolio_Master from Thomas's wealth-management workbook."

    def add_arguments(self, parser):
        parser.add_argument("--workbook", default=DEFAULT_WORKBOOK)
        parser.add_argument("--organization-id", type=int)
        parser.add_argument("--dry-run", action="store_true")

    def handle(self, *args, **options):
        path = Path(options["workbook"])
        if not path.exists():
            raise CommandError(f"Workbook not found: {path}")

        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise CommandError("openpyxl not installed.") from exc

        org = _resolve_organization(options.get("organization_id"))
        wb = load_workbook(path, data_only=True, read_only=False)
        if SHEET_NAME not in wb.sheetnames:
            raise CommandError(f"Sheet '{SHEET_NAME}' not found in workbook.")

        ws = wb[SHEET_NAME]
        rows = _rows_by_header(ws)

        stats = {
            "created": 0, "updated": 0,
            "skipped_no_id": 0, "skipped_explanation_row": 0,
            "linked_parent": 0,
        }
        errors: list[str] = []

        with transaction.atomic():
            # ── Pass 1: upsert ──
            for row in rows:
                portfolio_id = _text(row.get("Portfolio_ID"))
                if not portfolio_id:
                    stats["skipped_no_id"] += 1
                    continue
                if not portfolio_id.startswith("PORT_"):
                    stats["skipped_explanation_row"] += 1
                    continue
                portfolio_name = _text(row.get("Portfolio_Name"))
                portfolio_type_value = _text(row.get("Portfolio_Type"))
                if not portfolio_name and not portfolio_type_value:
                    stats["skipped_explanation_row"] += 1
                    continue
                if " " in portfolio_id or "=" in portfolio_id:
                    stats["skipped_explanation_row"] += 1
                    continue

                defaults = self._build_defaults(row)
                _, created = Portfolio.objects.update_or_create(
                    organization=org,
                    portfolio_id=portfolio_id,
                    defaults=defaults,
                )
                stats["created" if created else "updated"] += 1

            # ── Pass 2: resolve parent FKs ──
            lookup = {
                p.portfolio_id: p
                for p in Portfolio.objects.filter(organization=org)
            }
            for portfolio in lookup.values():
                code = portfolio.parent_portfolio_workbook_id
                if not code:
                    if portfolio.parent_id is not None:
                        portfolio.parent = None
                        portfolio.save(update_fields=["parent"])
                    continue
                parent = lookup.get(code)
                if parent is None:
                    errors.append(
                        f"Portfolio '{portfolio.portfolio_id}' references "
                        f"unknown parent '{code}'."
                    )
                    continue
                if parent.id == portfolio.id:
                    errors.append(
                        f"Portfolio '{portfolio.portfolio_id}' is its own parent — skipping."
                    )
                    continue
                if portfolio.parent_id != parent.id:
                    portfolio.parent = parent
                    portfolio.save(update_fields=["parent"])
                stats["linked_parent"] += 1

            if options["dry_run"]:
                transaction.set_rollback(True)

        prefix = "DRY RUN parsed" if options["dry_run"] else "Imported"
        self.stdout.write(self.style.SUCCESS(
            f"{prefix} Portfolio rows for org {org.id} ({org.name})"
        ))
        for k, v in stats.items():
            self.stdout.write(f"  {k}: {v}")
        if errors:
            self.stdout.write(self.style.WARNING(f"{len(errors)} warning(s):"))
            for e in errors[:20]:
                self.stdout.write(f"  - {e}")
            if len(errors) > 20:
                self.stdout.write(f"  ... and {len(errors) - 20} more")

    # ─────────────── per-row mapping ─────────────── #

    def _build_defaults(self, row: dict[str, Any]) -> dict:
        defaults: dict[str, Any] = {}
        metadata: dict[str, Any] = {}

        for column, value in row.items():
            if column == "_row":
                continue
            if column in COLUMN_TO_FIELD:
                field = COLUMN_TO_FIELD[column]
                defaults[field] = self._coerce(field, value)
            else:
                if value not in (None, ""):
                    metadata[column] = _json_safe(value)

        # Normalise enum-ish strings.
        for key in ("portfolio_type", "owner_type", "status"):
            if defaults.get(key):
                defaults[key] = defaults[key].upper()
        if defaults.get("base_currency"):
            defaults["base_currency"] = defaults["base_currency"].upper()[:3]
        if defaults.get("reporting_currency"):
            defaults["reporting_currency"] = defaults["reporting_currency"].upper()[:3]

        defaults["workbook_metadata"] = metadata
        return defaults

    def _coerce(self, field: str, value: Any) -> Any:
        if field in DATE_FIELDS:
            return _date(value)
        if field in BOOL_FIELDS:
            return _bool(value, default=False)
        return _text(value)


# ────────────────────────── helpers ────────────────────────── #


def _resolve_organization(organization_id: int | None) -> Organization:
    if organization_id:
        try:
            return Organization.objects.get(id=organization_id)
        except Organization.DoesNotExist as exc:
            raise CommandError(f"Organization id {organization_id} does not exist.") from exc

    orgs = list(Organization.objects.all()[:2])
    if not orgs:
        raise CommandError("No Organization exists. Pass --organization-id.")
    if len(orgs) > 1:
        raise CommandError("Multiple organizations exist. Re-run with --organization-id.")
    return orgs[0]


def _rows_by_header(ws) -> list[dict[str, Any]]:
    headers = [_text(ws.cell(1, col).value) for col in range(1, ws.max_column + 1)]
    rows: list[dict[str, Any]] = []
    for row_no in range(2, ws.max_row + 1):
        row: dict[str, Any] = {"_row": row_no}
        has_value = False
        for idx, header in enumerate(headers, start=1):
            value = _value(ws.cell(row_no, idx).value)
            if value not in ("", None):
                has_value = True
            if header:
                row[header] = value
        if has_value:
            rows.append(row)
    return rows


def _value(value: Any) -> Any:
    if isinstance(value, str):
        return value.strip()
    return value


def _text(value: Any) -> str:
    value = _value(value)
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, Decimal):
        return str(value)
    return str(value).strip()


def _bool(value: Any, *, default: bool) -> bool:
    text = _text(value).lower()
    if not text:
        return default
    if text in {"yes", "y", "true", "1", "active"}:
        return True
    if text in {"no", "n", "false", "0", "inactive"}:
        return False
    return default


def _date(value: Any):
    if value in ("", None):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    try:
        return datetime.strptime(_text(value), "%Y-%m-%d").date()
    except ValueError:
        return None


def _json_safe(value: Any) -> Any:
    if isinstance(value, dict):
        return {str(k): _json_safe(v) for k, v in value.items() if not str(k).startswith("_")}
    if isinstance(value, list):
        return [_json_safe(v) for v in value]
    if isinstance(value, tuple):
        return [_json_safe(v) for v in value]
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, Decimal):
        return str(value)
    return value
