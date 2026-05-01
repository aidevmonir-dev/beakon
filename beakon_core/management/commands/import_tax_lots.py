"""Import the workbook tab `17_Tax_Lot_Master` into the TaxLot master table.

Usage:
    python manage.py import_tax_lots [--workbook PATH] [--organization-id ID] [--dry-run]

Idempotent: re-running the command updates existing rows in place keyed on
``(organization, tax_lot_id)``. Missing ``Account_No`` cells leave the FK to
``Account`` null but still preserve the raw value in ``account_no`` for audit.
"""
from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from beakon_core.models import Account, TaxLot
from organizations.models import Organization


DEFAULT_WORKBOOK = r"D:\Thomas\2026 04 17-DRAFT-CoA-Wealth management v2.xlsx"
SHEET_NAME = "17_Tax_Lot_Master"


# Workbook column → model field. Anything not in this map is preserved into
# `workbook_metadata` so the round-trip stays lossless.
DATE_FIELDS = {"lot_open_date", "acquisition_trade_date", "settlement_date", "disposal_date"}
DECIMAL_FIELDS = {
    "original_quantity", "remaining_quantity", "acquisition_price_per_unit",
    "acquisition_fx_rate_to_reporting", "acquisition_cost_transaction_ccy",
    "acquisition_cost_reporting_ccy", "disposed_quantity",
    "cumulative_disposed_quantity", "remaining_cost_reporting_ccy",
    "realized_gain_loss_reporting_ccy",
}
BOOL_FIELDS = {"wash_sale_flag", "corporate_action_adjusted_flag", "active_flag"}

COLUMN_TO_FIELD: dict[str, str] = {
    "Tax_Lot_ID": "tax_lot_id",
    "Instrument_ID": "instrument_code",
    "Portfolio_ID": "portfolio_code",
    "Custodian_ID": "custodian_code",
    "Account_No": "account_no",
    "Lot_Open_Date": "lot_open_date",
    "Acquisition_Trade_Date": "acquisition_trade_date",
    "Settlement_Date": "settlement_date",
    "Original_Quantity": "original_quantity",
    "Remaining_Quantity": "remaining_quantity",
    "Unit_Of_Measure": "unit_of_measure",
    "Acquisition_Price_Per_Unit": "acquisition_price_per_unit",
    "Acquisition_Currency": "acquisition_currency",
    "Acquisition_FX_Rate_to_Reporting": "acquisition_fx_rate_to_reporting",
    "Acquisition_Cost_Transaction_Ccy": "acquisition_cost_transaction_ccy",
    "Acquisition_Cost_Reporting_Ccy": "acquisition_cost_reporting_ccy",
    "Cost_Basis_Method": "cost_basis_method",
    "Lot_Status": "lot_status",
    "Disposal_Date": "disposal_date",
    "Disposed_Quantity": "disposed_quantity",
    "Cumulative_Disposed_Quantity": "cumulative_disposed_quantity",
    "Remaining_Cost_Reporting_Ccy": "remaining_cost_reporting_ccy",
    "Realized_Gain_Loss_Reporting_Ccy": "realized_gain_loss_reporting_ccy",
    "Wash_Sale_Flag": "wash_sale_flag",
    "Corporate_Action_Adjusted_Flag": "corporate_action_adjusted_flag",
    "Source_Transaction_Reference": "source_transaction_reference",
    "Source_Document_Reference": "source_document_reference",
    "Active_Flag": "active_flag",
    "Notes": "notes",
}


class Command(BaseCommand):
    help = "Import tab 17_Tax_Lot_Master from Thomas's wealth-management workbook."

    def add_arguments(self, parser):
        parser.add_argument("--workbook", default=DEFAULT_WORKBOOK)
        parser.add_argument("--organization-id", type=int)
        parser.add_argument("--dry-run", action="store_true")

    def handle(self, *args, **options):
        path = Path(options["workbook"])
        if not path.exists():
            raise CommandError(f"Workbook not found: {path}")

        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise CommandError("openpyxl not installed.") from exc

        org = _resolve_organization(options.get("organization_id"))
        wb = load_workbook(path, data_only=True, read_only=False)
        if SHEET_NAME not in wb.sheetnames:
            raise CommandError(f"Sheet '{SHEET_NAME}' not found in workbook.")

        ws = wb[SHEET_NAME]
        rows = _rows_by_header(ws)

        # Pre-build account lookup so the import is one query, not N+1.
        account_lookup = {
            a.code: a for a in Account.objects.filter(organization=org)
        }

        stats = {
            "created": 0, "updated": 0, "skipped_no_id": 0,
            "skipped_explanation_row": 0, "linked_account": 0,
        }
        errors: list[str] = []

        with transaction.atomic():
            for row in rows:
                tax_lot_id = _text(row.get("Tax_Lot_ID"))
                if not tax_lot_id:
                    stats["skipped_no_id"] += 1
                    continue
                # Tabs in this workbook contain explanation rows below the
                # data. Real lots follow the `TLOT_…` workbook convention.
                if not tax_lot_id.startswith("TLOT_"):
                    stats["skipped_explanation_row"] += 1
                    continue

                defaults = self._build_defaults(row, account_lookup, errors, stats)
                _, created = TaxLot.objects.update_or_create(
                    organization=org,
                    tax_lot_id=tax_lot_id,
                    defaults=defaults,
                )
                stats["created" if created else "updated"] += 1

            if options["dry_run"]:
                transaction.set_rollback(True)

        prefix = "DRY RUN parsed" if options["dry_run"] else "Imported"
        self.stdout.write(self.style.SUCCESS(
            f"{prefix} TaxLot rows for org {org.id} ({org.name})"
        ))
        for k, v in stats.items():
            self.stdout.write(f"  {k}: {v}")
        if errors:
            self.stdout.write(self.style.WARNING(f"{len(errors)} warning(s):"))
            for e in errors[:20]:
                self.stdout.write(f"  - {e}")
            if len(errors) > 20:
                self.stdout.write(f"  ... and {len(errors) - 20} more")

    # ─────────────── per-row mapping ─────────────── #

    def _build_defaults(self, row: dict[str, Any], account_lookup, errors, stats) -> dict:
        defaults: dict[str, Any] = {}
        metadata: dict[str, Any] = {}

        for column, value in row.items():
            if column == "_row":
                continue
            if column in COLUMN_TO_FIELD:
                field = COLUMN_TO_FIELD[column]
                defaults[field] = self._coerce(field, value)
            else:
                # Unknown column → preserve into metadata for round-trip safety.
                if value not in (None, ""):
                    metadata[column] = _json_safe(value)

        # Promote known choices to canonical form.
        if "cost_basis_method" in defaults and defaults["cost_basis_method"]:
            defaults["cost_basis_method"] = defaults["cost_basis_method"].upper()
        if "lot_status" in defaults and defaults["lot_status"]:
            defaults["lot_status"] = defaults["lot_status"].upper()
        if "acquisition_currency" in defaults and defaults["acquisition_currency"]:
            defaults["acquisition_currency"] = defaults["acquisition_currency"].upper()[:3]

        # Resolve Account FK from Account_No (if present).
        account_no = _text(defaults.get("account_no"))
        if account_no:
            account = account_lookup.get(account_no)
            if account is not None:
                defaults["account"] = account
                stats["linked_account"] += 1
            else:
                errors.append(
                    f"Tax lot '{defaults.get('tax_lot_id')}' references "
                    f"unknown Account_No '{account_no}'."
                )

        defaults["workbook_metadata"] = metadata
        return defaults

    def _coerce(self, field: str, value: Any) -> Any:
        if field in DATE_FIELDS:
            return _date(value)
        if field in DECIMAL_FIELDS:
            return _decimal(value, default=None)
        if field in BOOL_FIELDS:
            return _bool(value, default=False)
        return _text(value)


# ────────────────────────── helpers (mirrored from import_phase1_coa) ────────────────────────── #


def _resolve_organization(organization_id: int | None) -> Organization:
    if organization_id:
        try:
            return Organization.objects.get(id=organization_id)
        except Organization.DoesNotExist as exc:
            raise CommandError(f"Organization id {organization_id} does not exist.") from exc

    orgs = list(Organization.objects.all()[:2])
    if not orgs:
        raise CommandError(
            "No Organization exists. Create one first or pass --organization-id."
        )
    if len(orgs) > 1:
        raise CommandError(
            "Multiple organizations exist. Re-run with --organization-id <id>."
        )
    return orgs[0]


def _rows_by_header(ws) -> list[dict[str, Any]]:
    headers = [_text(ws.cell(1, col).value) for col in range(1, ws.max_column + 1)]
    rows: list[dict[str, Any]] = []
    for row_no in range(2, ws.max_row + 1):
        row: dict[str, Any] = {"_row": row_no}
        has_value = False
        for idx, header in enumerate(headers, start=1):
            value = _value(ws.cell(row_no, idx).value)
            if value not in ("", None):
                has_value = True
            if header:
                row[header] = value
        if has_value:
            rows.append(row)
    return rows


def _value(value: Any) -> Any:
    if isinstance(value, str):
        return value.strip()
    return value


def _text(value: Any) -> str:
    value = _value(value)
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, Decimal):
        return str(value)
    return str(value).strip()


def _bool(value: Any, *, default: bool) -> bool:
    text = _text(value).lower()
    if not text:
        return default
    if text in {"yes", "y", "true", "1", "active"}:
        return True
    if text in {"no", "n", "false", "0", "inactive"}:
        return False
    return default


def _decimal(value: Any, *, default):
    if value in ("", None):
        return default
    try:
        return Decimal(str(value))
    except Exception:
        return default


def _date(value: Any):
    if value in ("", None):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    try:
        return datetime.strptime(_text(value), "%Y-%m-%d").date()
    except ValueError:
        return None


def _json_safe(value: Any) -> Any:
    if isinstance(value, dict):
        return {str(k): _json_safe(v) for k, v in value.items() if not str(k).startswith("_")}
    if isinstance(value, list):
        return [_json_safe(v) for v in value]
    if isinstance(value, tuple):
        return [_json_safe(v) for v in value]
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, Decimal):
        return str(value)
    return value
