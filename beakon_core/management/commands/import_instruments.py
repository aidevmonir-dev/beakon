"""Import the workbook tab `08 Instrument Master` into the Instrument table.

Usage:
    python manage.py import_instruments [--workbook PATH] [--organization-id ID] [--dry-run]

Idempotent: re-running updates existing rows in place keyed on
``(organization, instrument_id)``. Resolves six ``Default_*_Account`` codes
to ``Account`` foreign keys and the ``Loan_ID`` column to a ``Loan`` FK
when matching workbook IDs exist.
"""
from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from beakon_core.models import Account, Instrument, Loan
from organizations.models import Organization


DEFAULT_WORKBOOK = r"D:\Thomas\2026 04 17-DRAFT-CoA-Wealth management v2.xlsx"
SHEET_NAME = "08 Instrument Master"

DATE_FIELDS = {
    "inception_date", "maturity_date", "effective_from", "effective_to",
}
BOOL_FIELDS = {
    "related_party_flag", "commitment_flag", "loan_linked_flag",
    "tax_lot_required", "fx_exposure_flag", "esg_or_restriction_flag",
}

COLUMN_TO_FIELD: dict[str, str] = {
    "Instrument_ID": "instrument_id",
    "Instrument_Name": "instrument_name",
    "Instrument_Type": "instrument_type",
    "Quoted_Unquoted_Flag": "quoted_unquoted_flag",
    "Asset_Class_Code": "asset_class_code",
    "Strategy_Code": "strategy_code",
    "Portfolio_Default": "portfolio_default",
    "Custodian_Default": "custodian_default",
    "Issuer_or_Counterparty_Code": "issuer_or_counterparty_code",
    "Related_Party_Flag": "related_party_flag",
    "ISIN_or_Ticker": "isin_or_ticker",
    "Internal_Reference": "internal_reference",
    "Currency": "currency",
    "Jurisdiction_Code": "jurisdiction_code",
    "Domicile_Code": "domicile_code",
    "Commitment_Flag": "commitment_flag",
    "Commitment_Code": "commitment_code",
    "Loan_Linked_Flag": "loan_linked_flag",
    "Loan_ID": "loan_workbook_id",
    "Tax_Lot_Required": "tax_lot_required",
    "Income_Type": "income_type",
    "Income_Frequency": "income_frequency",
    "Valuation_Method": "valuation_method",
    "Price_Source": "price_source",
    "FX_Exposure_Flag": "fx_exposure_flag",
    "Impairment_Method": "impairment_method",
    "ESG_or_Restriction_Flag": "esg_or_restriction_flag",
    "Restriction_Type_Code": "restriction_type_code",
    "Inception_Date": "inception_date",
    "Maturity_Date": "maturity_date",
    "Settlement_Cycle": "settlement_cycle",
    "Day_Count_Convention": "day_count_convention",
    "Performance_Group": "performance_group",
    "Report_Category_Code": "report_category_code",
    "Default_Principal_Account": "default_principal_account_code",
    "Default_Income_Account": "default_income_account_code",
    "Default_Expense_Account": "default_expense_account_code",
    "Default_Realized_GL_Account": "default_realized_gl_account_code",
    "Default_Unrealized_GL_Account": "default_unrealized_gl_account_code",
    "Default_FX_GL_Account": "default_fx_gl_account_code",
    "Status": "status",
    "Effective_From": "effective_from",
    "Effective_To": "effective_to",
    "Notes": "notes",
}

ACCOUNT_FK_FIELDS = [
    ("default_principal_account_code", "default_principal_account", "principal"),
    ("default_income_account_code", "default_income_account", "income"),
    ("default_expense_account_code", "default_expense_account", "expense"),
    ("default_realized_gl_account_code", "default_realized_gl_account", "realized_gl"),
    ("default_unrealized_gl_account_code", "default_unrealized_gl_account", "unrealized_gl"),
    ("default_fx_gl_account_code", "default_fx_gl_account", "fx_gl"),
]


class Command(BaseCommand):
    help = "Import tab 08 Instrument Master from Thomas's wealth-management workbook."

    def add_arguments(self, parser):
        parser.add_argument("--workbook", default=DEFAULT_WORKBOOK)
        parser.add_argument("--organization-id", type=int)
        parser.add_argument("--dry-run", action="store_true")

    def handle(self, *args, **options):
        path = Path(options["workbook"])
        if not path.exists():
            raise CommandError(f"Workbook not found: {path}")

        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise CommandError("openpyxl not installed.") from exc

        org = _resolve_organization(options.get("organization_id"))
        wb = load_workbook(path, data_only=True, read_only=False)
        if SHEET_NAME not in wb.sheetnames:
            raise CommandError(f"Sheet '{SHEET_NAME}' not found in workbook.")

        ws = wb[SHEET_NAME]
        rows = _rows_by_header(ws)

        account_lookup = {a.code: a for a in Account.objects.filter(organization=org)}
        loan_lookup = {l.loan_id: l for l in Loan.objects.filter(organization=org)}

        stats = {
            "created": 0, "updated": 0,
            "skipped_no_id": 0, "skipped_explanation_row": 0,
            "linked_loan": 0,
            "linked_principal": 0, "linked_income": 0,
            "linked_expense": 0, "linked_realized_gl": 0,
            "linked_unrealized_gl": 0, "linked_fx_gl": 0,
        }
        errors: list[str] = []

        with transaction.atomic():
            for row in rows:
                instrument_id = _text(row.get("Instrument_ID"))
                if not instrument_id:
                    stats["skipped_no_id"] += 1
                    continue
                # Real instruments use the INS_… prefix in this workbook.
                if not (instrument_id.startswith("INS_") or instrument_id.startswith("INST_")):
                    stats["skipped_explanation_row"] += 1
                    continue
                # Stubs/explanation rows have an ID-like prefix but no other
                # meaningful data. Require at least an instrument_name or
                # instrument_type to count as a real row.
                instrument_name = _text(row.get("Instrument_Name"))
                instrument_type_value = _text(row.get("Instrument_Type"))
                if not instrument_name and not instrument_type_value:
                    stats["skipped_explanation_row"] += 1
                    continue
                # Reject IDs that contain whitespace or '=' — those are
                # explanation-row text (e.g. "INST_EQ_MSFT = Microsoft …").
                if " " in instrument_id or "=" in instrument_id:
                    stats["skipped_explanation_row"] += 1
                    continue

                defaults = self._build_defaults(
                    row, account_lookup, loan_lookup, errors, stats,
                )
                _, created = Instrument.objects.update_or_create(
                    organization=org,
                    instrument_id=instrument_id,
                    defaults=defaults,
                )
                stats["created" if created else "updated"] += 1

            if options["dry_run"]:
                transaction.set_rollback(True)

        prefix = "DRY RUN parsed" if options["dry_run"] else "Imported"
        self.stdout.write(self.style.SUCCESS(
            f"{prefix} Instrument rows for org {org.id} ({org.name})"
        ))
        for k, v in stats.items():
            self.stdout.write(f"  {k}: {v}")
        if errors:
            self.stdout.write(self.style.WARNING(f"{len(errors)} warning(s):"))
            for e in errors[:20]:
                self.stdout.write(f"  - {e}")
            if len(errors) > 20:
                self.stdout.write(f"  ... and {len(errors) - 20} more")

    # ─────────────── per-row mapping ─────────────── #

    def _build_defaults(self, row: dict[str, Any], account_lookup, loan_lookup,
                        errors, stats) -> dict:
        defaults: dict[str, Any] = {}
        metadata: dict[str, Any] = {}

        for column, value in row.items():
            if column == "_row":
                continue
            if column in COLUMN_TO_FIELD:
                field = COLUMN_TO_FIELD[column]
                defaults[field] = self._coerce(field, value)
            else:
                if value not in (None, ""):
                    metadata[column] = _json_safe(value)

        # Normalise enum-ish strings.
        for key in ("instrument_type", "quoted_unquoted_flag", "income_type",
                    "income_frequency", "valuation_method", "price_source",
                    "impairment_method", "settlement_cycle", "status"):
            if defaults.get(key):
                defaults[key] = defaults[key].upper()

        # Resolve Loan FK from workbook code.
        loan_workbook_id = _text(defaults.get("loan_workbook_id"))
        if loan_workbook_id:
            loan = loan_lookup.get(loan_workbook_id)
            if loan is not None:
                defaults["loan"] = loan
                stats["linked_loan"] += 1
            else:
                errors.append(
                    f"Instrument '{defaults.get('instrument_id')}' references "
                    f"unknown Loan_ID '{loan_workbook_id}'."
                )

        # Resolve default GL account FK columns.
        for code_field, fk_field, stat_key in ACCOUNT_FK_FIELDS:
            code = _text(defaults.get(code_field))
            if not code or not code.replace(".", "").isdigit():
                continue
            account = account_lookup.get(code)
            if account is not None:
                defaults[fk_field] = account
                stats[f"linked_{stat_key}"] += 1
            else:
                errors.append(
                    f"Instrument '{defaults.get('instrument_id')}' references "
                    f"unknown {code_field} '{code}'."
                )

        defaults["workbook_metadata"] = metadata
        return defaults

    def _coerce(self, field: str, value: Any) -> Any:
        if field in DATE_FIELDS:
            return _date(value)
        if field in BOOL_FIELDS:
            return _bool(value, default=False)
        return _text(value)


# ────────────────────────── helpers ────────────────────────── #


def _resolve_organization(organization_id: int | None) -> Organization:
    if organization_id:
        try:
            return Organization.objects.get(id=organization_id)
        except Organization.DoesNotExist as exc:
            raise CommandError(f"Organization id {organization_id} does not exist.") from exc

    orgs = list(Organization.objects.all()[:2])
    if not orgs:
        raise CommandError("No Organization exists. Pass --organization-id.")
    if len(orgs) > 1:
        raise CommandError("Multiple organizations exist. Re-run with --organization-id.")
    return orgs[0]


def _rows_by_header(ws) -> list[dict[str, Any]]:
    headers = [_text(ws.cell(1, col).value) for col in range(1, ws.max_column + 1)]
    rows: list[dict[str, Any]] = []
    for row_no in range(2, ws.max_row + 1):
        row: dict[str, Any] = {"_row": row_no}
        has_value = False
        for idx, header in enumerate(headers, start=1):
            value = _value(ws.cell(row_no, idx).value)
            if value not in ("", None):
                has_value = True
            if header:
                row[header] = value
        if has_value:
            rows.append(row)
    return rows


def _value(value: Any) -> Any:
    if isinstance(value, str):
        return value.strip()
    return value


def _text(value: Any) -> str:
    value = _value(value)
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, Decimal):
        return str(value)
    return str(value).strip()


def _bool(value: Any, *, default: bool) -> bool:
    text = _text(value).lower()
    if not text:
        return default
    if text in {"yes", "y", "true", "1", "active"}:
        return True
    if text in {"no", "n", "false", "0", "inactive"}:
        return False
    return default


def _date(value: Any):
    if value in ("", None):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    try:
        return datetime.strptime(_text(value), "%Y-%m-%d").date()
    except ValueError:
        return None


def _json_safe(value: Any) -> Any:
    if isinstance(value, dict):
        return {str(k): _json_safe(v) for k, v in value.items() if not str(k).startswith("_")}
    if isinstance(value, list):
        return [_json_safe(v) for v in value]
    if isinstance(value, tuple):
        return [_json_safe(v) for v in value]
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, Decimal):
        return str(value)
    return value
