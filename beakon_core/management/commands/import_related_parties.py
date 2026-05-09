"""Import the workbook tab `11_Related_Party_Master` into the RelatedParty table.

Usage:
    python manage.py import_related_parties [--workbook PATH] [--organization-id ID] [--dry-run]

Idempotent: re-running updates existing rows in place keyed on
``(organization, related_party_id)``. Resolves ``Default_Portfolio_Code`` to a
``Portfolio`` foreign key when the code matches.
"""
from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from beakon_core.models import Portfolio, RelatedParty
from organizations.models import Organization


DEFAULT_WORKBOOK = r"D:\Thomas\2026 04 17-DRAFT-CoA-Wealth management v2.xlsx"
SHEET_NAME = "11_Related_Party_Master"

DATE_FIELDS = {"related_party_since", "related_party_until"}
DECIMAL_FIELDS = {"ownership_percent"}
BOOL_FIELDS = {
    "control_flag", "beneficiary_flag", "settlor_flag", "protector_flag",
    "director_flag", "signer_flag", "active_flag",
    "loan_eligible_flag", "distribution_eligible_flag",
    "capital_contribution_eligible_flag",
    "expense_allocation_eligible_flag", "family_expense_eligible_flag",
    "net_worth_inclusion_flag",
    "source_document_required_flag", "approval_required_flag",
}

COLUMN_TO_FIELD: dict[str, str] = {
    "Related_Party_ID": "related_party_id",
    "Related_Party_Name": "related_party_name",
    "Short_Name": "short_name",
    "Related_Party_Type": "related_party_type",
    "Related_Party_Subtype": "related_party_subtype",
    "Party_Form": "party_form",
    "Relationship_To_Client": "relationship_to_client",
    "Ownership_Percent": "ownership_percent",
    "Control_Flag": "control_flag",
    "Beneficiary_Flag": "beneficiary_flag",
    "Settlor_Flag": "settlor_flag",
    "Protector_Flag": "protector_flag",
    "Director_Flag": "director_flag",
    "Signer_Flag": "signer_flag",
    "Status": "status",
    "Active_Flag": "active_flag",
    "Country_Code": "country_code",
    "Jurisdiction_Code": "jurisdiction_code",
    "Base_Currency": "base_currency",
    "Reporting_Currency": "reporting_currency",
    "Tax_Residence_Country": "tax_residence_country",
    "Related_Party_Since": "related_party_since",
    "Related_Party_Until": "related_party_until",
    "Loan_Eligible_Flag": "loan_eligible_flag",
    "Distribution_Eligible_Flag": "distribution_eligible_flag",
    "Capital_Contribution_Eligible_Flag": "capital_contribution_eligible_flag",
    "Expense_Allocation_Eligible_Flag": "expense_allocation_eligible_flag",
    "Family_Expense_Eligible_Flag": "family_expense_eligible_flag",
    "Net_Worth_Inclusion_Flag": "net_worth_inclusion_flag",
    "Default_Portfolio_Code": "default_portfolio_code",
    "Default_Property_Code": "default_property_code",
    "Default_Bank_Reference": "default_bank_reference",
    "Source_Document_Required_Flag": "source_document_required_flag",
    "Approval_Required_Flag": "approval_required_flag",
    "Notes": "notes",
}


class Command(BaseCommand):
    help = "Import tab 11_Related_Party_Master from Thomas's wealth-management workbook."

    def add_arguments(self, parser):
        parser.add_argument("--workbook", default=DEFAULT_WORKBOOK)
        parser.add_argument("--organization-id", type=int)
        parser.add_argument("--dry-run", action="store_true")

    def handle(self, *args, **options):
        path = Path(options["workbook"])
        if not path.exists():
            raise CommandError(f"Workbook not found: {path}")

        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise CommandError("openpyxl not installed.") from exc

        org = _resolve_organization(options.get("organization_id"))
        wb = load_workbook(path, data_only=True, read_only=False)
        if SHEET_NAME not in wb.sheetnames:
            raise CommandError(f"Sheet '{SHEET_NAME}' not found in workbook.")

        ws = wb[SHEET_NAME]
        rows = _rows_by_header(ws)

        portfolio_lookup = {
            p.portfolio_id: p for p in Portfolio.objects.filter(organization=org)
        }

        stats = {
            "created": 0, "updated": 0,
            "skipped_no_id": 0, "skipped_explanation_row": 0,
            "linked_default_portfolio": 0,
        }
        errors: list[str] = []

        with transaction.atomic():
            for row in rows:
                rp_id = _text(row.get("Related_Party_ID"))
                if not rp_id:
                    stats["skipped_no_id"] += 1
                    continue
                # Real related parties use RP_* in this workbook draft.
                if not rp_id.startswith("RP_"):
                    stats["skipped_explanation_row"] += 1
                    continue
                rp_name = _text(row.get("Related_Party_Name"))
                rp_type = _text(row.get("Related_Party_Type"))
                if not rp_name and not rp_type:
                    stats["skipped_explanation_row"] += 1
                    continue
                if " " in rp_id or "=" in rp_id:
                    stats["skipped_explanation_row"] += 1
                    continue

                defaults = self._build_defaults(row, portfolio_lookup, errors, stats)
                _, created = RelatedParty.objects.update_or_create(
                    organization=org,
                    related_party_id=rp_id,
                    defaults=defaults,
                )
                stats["created" if created else "updated"] += 1

            if options["dry_run"]:
                transaction.set_rollback(True)

        prefix = "DRY RUN parsed" if options["dry_run"] else "Imported"
        self.stdout.write(self.style.SUCCESS(
            f"{prefix} RelatedParty rows for org {org.id} ({org.name})"
        ))
        for k, v in stats.items():
            self.stdout.write(f"  {k}: {v}")
        if errors:
            self.stdout.write(self.style.WARNING(f"{len(errors)} warning(s):"))
            for e in errors[:20]:
                self.stdout.write(f"  - {e}")
            if len(errors) > 20:
                self.stdout.write(f"  ... and {len(errors) - 20} more")

    # ─────────────── per-row mapping ─────────────── #

    def _build_defaults(self, row: dict[str, Any], portfolio_lookup, errors, stats) -> dict:
        defaults: dict[str, Any] = {}
        metadata: dict[str, Any] = {}

        for column, value in row.items():
            if column == "_row":
                continue
            if column in COLUMN_TO_FIELD:
                field = COLUMN_TO_FIELD[column]
                defaults[field] = self._coerce(field, value)
            else:
                if value not in (None, ""):
                    metadata[column] = _json_safe(value)

        for key in ("related_party_type", "party_form", "status"):
            if defaults.get(key):
                defaults[key] = defaults[key].upper()
        if defaults.get("base_currency"):
            defaults["base_currency"] = defaults["base_currency"].upper()[:3]
        if defaults.get("reporting_currency"):
            defaults["reporting_currency"] = defaults["reporting_currency"].upper()[:3]

        portfolio_code = _text(defaults.get("default_portfolio_code"))
        if portfolio_code:
            portfolio = portfolio_lookup.get(portfolio_code)
            if portfolio is not None:
                defaults["default_portfolio"] = portfolio
                stats["linked_default_portfolio"] += 1
            else:
                errors.append(
                    f"RelatedParty '{defaults.get('related_party_id')}' references "
                    f"unknown Default_Portfolio_Code '{portfolio_code}'."
                )

        defaults["workbook_metadata"] = metadata
        return defaults

    def _coerce(self, field: str, value: Any) -> Any:
        if field in DATE_FIELDS:
            return _date(value)
        if field in DECIMAL_FIELDS:
            return _decimal(value, default=None)
        if field in BOOL_FIELDS:
            return _bool(value, default=False)
        return _text(value)


# ────────────────────────── helpers ────────────────────────── #


def _resolve_organization(organization_id: int | None) -> Organization:
    if organization_id:
        try:
            return Organization.objects.get(id=organization_id)
        except Organization.DoesNotExist as exc:
            raise CommandError(f"Organization id {organization_id} does not exist.") from exc

    orgs = list(Organization.objects.all()[:2])
    if not orgs:
        raise CommandError("No Organization exists. Pass --organization-id.")
    if len(orgs) > 1:
        raise CommandError("Multiple organizations exist. Re-run with --organization-id.")
    return orgs[0]


def _rows_by_header(ws) -> list[dict[str, Any]]:
    headers = [_text(ws.cell(1, col).value) for col in range(1, ws.max_column + 1)]
    rows: list[dict[str, Any]] = []
    for row_no in range(2, ws.max_row + 1):
        row: dict[str, Any] = {"_row": row_no}
        has_value = False
        for idx, header in enumerate(headers, start=1):
            value = _value(ws.cell(row_no, idx).value)
            if value not in ("", None):
                has_value = True
            if header:
                row[header] = value
        if has_value:
            rows.append(row)
    return rows


def _value(value: Any) -> Any:
    if isinstance(value, str):
        return value.strip()
    return value


def _text(value: Any) -> str:
    value = _value(value)
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, Decimal):
        return str(value)
    return str(value).strip()


def _bool(value: Any, *, default: bool) -> bool:
    text = _text(value).lower()
    if not text:
        return default
    if text in {"yes", "y", "true", "1", "active"}:
        return True
    if text in {"no", "n", "false", "0", "inactive"}:
        return False
    return default


def _decimal(value: Any, *, default):
    if value in ("", None):
        return default
    try:
        return Decimal(str(value))
    except Exception:
        return default


def _date(value: Any):
    if value in ("", None):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    try:
        return datetime.strptime(_text(value), "%Y-%m-%d").date()
    except ValueError:
        return None


def _json_safe(value: Any) -> Any:
    if isinstance(value, dict):
        return {str(k): _json_safe(v) for k, v in value.items() if not str(k).startswith("_")}
    if isinstance(value, list):
        return [_json_safe(v) for v in value]
    if isinstance(value, tuple):
        return [_json_safe(v) for v in value]
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, Decimal):
        return str(value)
    return value
