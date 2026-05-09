"""Import workbook tab `12_Bank_Account_Master` (governed BankAccountMaster)."""
from __future__ import annotations

from pathlib import Path
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from beakon_core.models import (
    BankAccountMaster, Counterparty, Loan, Portfolio, RelatedParty,
)
from beakon_core.management.commands._master_import_helpers import (
    is_explanation_row, map_row, resolve_organization, rows_by_header, text,
)

DEFAULT_WORKBOOK = r"D:\Thomas\2026 04 17-DRAFT-CoA-Wealth management v2.xlsx"
SHEET_NAME = "12_Bank_Account_Master"

DATE_FIELDS = {"opening_date", "closing_date"}
DECIMAL_FIELDS = set()
BOOL_FIELDS = {
    "active_flag", "posting_allowed_flag", "restricted_flag",
    "interest_bearing_flag", "overdraft_allowed_flag",
    "approval_required_flag", "source_document_required_flag",
}

COLUMN_MAP = {
    "Bank_Account_ID": "bank_account_id",
    "Bank_Account_Name": "bank_account_name",
    "Short_Name": "short_name",
    "Account_Holder_Type": "account_holder_type",
    "Account_Holder_ID": "account_holder_id_code",
    "Bank_Counterparty_ID": "bank_counterparty_id_code",
    "Bank_Name": "bank_name",
    "Booking_Center": "booking_center",
    "IBAN_or_Account_No_Masked": "iban_or_account_no_masked",
    "SWIFT_BIC": "swift_bic",
    "Account_Currency": "account_currency",
    "Reporting_Currency": "reporting_currency",
    "Country_Code": "country_code",
    "Jurisdiction_Code": "jurisdiction_code",
    "Account_Type": "account_type",
    "Account_Subtype": "account_subtype",
    "Account_Purpose": "account_purpose",
    "Status": "status",
    "Active_Flag": "active_flag",
    "Posting_Allowed_Flag": "posting_allowed_flag",
    "Restricted_Flag": "restricted_flag",
    "Interest_Bearing_Flag": "interest_bearing_flag",
    "Overdraft_Allowed_Flag": "overdraft_allowed_flag",
    "Credit_Line_Linked_Loan_ID": "credit_line_linked_loan_id_code",
    "Default_Portfolio_Code": "default_portfolio_code",
    "Default_Related_Party_ID": "default_related_party_id_code",
    "Default_Counterparty_ID": "default_counterparty_id_code",
    "Opening_Date": "opening_date",
    "Closing_Date": "closing_date",
    "KYC_Status": "kyc_status",
    "Approval_Required_Flag": "approval_required_flag",
    "Source_Document_Required_Flag": "source_document_required_flag",
    "Notes": "notes",
}


class Command(BaseCommand):
    help = "Import tab 12_Bank_Account_Master."

    def add_arguments(self, parser):
        parser.add_argument("--workbook", default=DEFAULT_WORKBOOK)
        parser.add_argument("--organization-id", type=int)
        parser.add_argument("--dry-run", action="store_true")

    def handle(self, *args, **options):
        path = Path(options["workbook"])
        if not path.exists():
            raise CommandError(f"Workbook not found: {path}")
        from openpyxl import load_workbook
        org = resolve_organization(options.get("organization_id"))
        wb = load_workbook(path, data_only=True, read_only=False)
        if SHEET_NAME not in wb.sheetnames:
            raise CommandError(f"Sheet '{SHEET_NAME}' missing.")
        rows = rows_by_header(wb[SHEET_NAME])

        rp_lookup = {p.related_party_id: p for p in RelatedParty.objects.filter(organization=org)}
        cp_lookup = {p.counterparty_id: p for p in Counterparty.objects.filter(organization=org)}
        loan_lookup = {p.loan_id: p for p in Loan.objects.filter(organization=org)}
        port_lookup = {p.portfolio_id: p for p in Portfolio.objects.filter(organization=org)}

        stats = {
            "created": 0, "updated": 0,
            "skipped_no_id": 0, "skipped_explanation_row": 0,
            "linked_holder_rp": 0, "linked_bank_cp": 0,
            "linked_credit_line_loan": 0, "linked_default_portfolio": 0,
            "linked_default_rp": 0, "linked_default_cp": 0,
        }
        errors: list[str] = []

        with transaction.atomic():
            for row in rows:
                ba_id = text(row.get("Bank_Account_ID"))
                if not ba_id:
                    stats["skipped_no_id"] += 1
                    continue
                if is_explanation_row(ba_id, prefix="BANK_",
                                       name=text(row.get("Bank_Account_Name")),
                                       type_=text(row.get("Account_Type"))):
                    stats["skipped_explanation_row"] += 1
                    continue

                defaults, metadata = map_row(
                    row, COLUMN_MAP,
                    date_fields=DATE_FIELDS, decimal_fields=DECIMAL_FIELDS,
                    bool_fields=BOOL_FIELDS,
                )
                for k in ("account_holder_type", "account_type", "status"):
                    if defaults.get(k):
                        defaults[k] = defaults[k].upper()
                if defaults.get("account_currency"):
                    defaults["account_currency"] = defaults["account_currency"].upper()[:3]
                if defaults.get("reporting_currency"):
                    defaults["reporting_currency"] = defaults["reporting_currency"].upper()[:3]

                # FK resolutions
                def link(code_field, fk_field, lookup, stat_key, master_label):
                    code = text(defaults.get(code_field))
                    if not code:
                        return
                    target = lookup.get(code)
                    if target is not None:
                        defaults[fk_field] = target
                        stats[stat_key] += 1
                    else:
                        errors.append(
                            f"BankAccount '{ba_id}' references unknown {master_label} "
                            f"'{code}' in column {code_field}."
                        )

                link("account_holder_id_code", "account_holder_related_party",
                     rp_lookup, "linked_holder_rp", "RelatedParty")
                link("bank_counterparty_id_code", "bank_counterparty",
                     cp_lookup, "linked_bank_cp", "Counterparty")
                link("credit_line_linked_loan_id_code", "credit_line_linked_loan",
                     loan_lookup, "linked_credit_line_loan", "Loan")
                link("default_portfolio_code", "default_portfolio",
                     port_lookup, "linked_default_portfolio", "Portfolio")
                link("default_related_party_id_code", "default_related_party",
                     rp_lookup, "linked_default_rp", "RelatedParty")
                link("default_counterparty_id_code", "default_counterparty",
                     cp_lookup, "linked_default_cp", "Counterparty")

                defaults["workbook_metadata"] = metadata
                _, created = BankAccountMaster.objects.update_or_create(
                    organization=org, bank_account_id=ba_id, defaults=defaults,
                )
                stats["created" if created else "updated"] += 1

            if options["dry_run"]:
                transaction.set_rollback(True)

        prefix = "DRY RUN parsed" if options["dry_run"] else "Imported"
        self.stdout.write(self.style.SUCCESS(f"{prefix} BankAccountMaster rows for org {org.id} ({org.name})"))
        for k, v in stats.items():
            self.stdout.write(f"  {k}: {v}")
        if errors:
            self.stdout.write(self.style.WARNING(f"{len(errors)} warning(s):"))
            for e in errors[:20]:
                self.stdout.write(f"  - {e}")
