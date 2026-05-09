"""Import workbook tab `16_Policy_Master`."""
from __future__ import annotations

from pathlib import Path
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from beakon_core.models import (
    Counterparty, Policy, Portfolio, Property, RelatedParty,
)
from beakon_core.management.commands._master_import_helpers import (
    clip_short_codes, is_explanation_row, map_row, resolve_organization,
    rows_by_header, text,
)

DEFAULT_WORKBOOK = r"D:\Thomas\2026 04 17-DRAFT-CoA-Wealth management v2.xlsx"
SHEET_NAME = "16_Policy_Master"

DATE_FIELDS = {"inception_date", "expiry_date", "renewal_date"}
DECIMAL_FIELDS = {"coverage_amount", "deductible_amount"}
BOOL_FIELDS = {
    "active_flag", "posting_allowed_flag",
    "approval_required_flag", "source_document_required_flag",
    "investment_linked_flag", "claim_eligible_flag", "premium_payable_flag",
}

COLUMN_MAP = {
    "Policy_ID": "policy_id",
    "Policy_Name": "policy_name",
    "Short_Name": "short_name",
    "Policy_Type": "policy_type",
    "Policy_Subtype": "policy_subtype",
    "Policy_Owner_Type": "policy_owner_type",
    "Policy_Owner_ID": "policy_owner_id_code",
    "Primary_Related_Party_ID": "primary_related_party_id_code",
    "Insured_Party_ID": "insured_party_id_code",
    "Insurer_Counterparty_ID": "insurer_counterparty_id_code",
    "Policy_Number_Masked": "policy_number_masked",
    "Country_Code": "country_code",
    "Jurisdiction_Code": "jurisdiction_code",
    "Policy_Currency": "policy_currency",
    "Reporting_Currency": "reporting_currency",
    "Inception_Date": "inception_date",
    "Expiry_Date": "expiry_date",
    "Renewal_Date": "renewal_date",
    "Premium_Frequency": "premium_frequency",
    "Coverage_Amount": "coverage_amount",
    "Deductible_Amount": "deductible_amount",
    "Investment_Linked_Flag": "investment_linked_flag",
    "Linked_Portfolio_ID": "linked_portfolio_id_code",
    "Linked_Property_ID": "linked_property_id_code",
    "Linked_Beneficiary_ID": "linked_beneficiary_id_code",
    "Claim_Eligible_Flag": "claim_eligible_flag",
    "Premium_Payable_Flag": "premium_payable_flag",
    "Status": "status",
    "Active_Flag": "active_flag",
    "Posting_Allowed_Flag": "posting_allowed_flag",
    "Approval_Required_Flag": "approval_required_flag",
    "Source_Document_Required_Flag": "source_document_required_flag",
    "Notes": "notes",
}


class Command(BaseCommand):
    help = "Import tab 16_Policy_Master."

    def add_arguments(self, parser):
        parser.add_argument("--workbook", default=DEFAULT_WORKBOOK)
        parser.add_argument("--organization-id", type=int)
        parser.add_argument("--dry-run", action="store_true")

    def handle(self, *args, **options):
        path = Path(options["workbook"])
        if not path.exists():
            raise CommandError(f"Workbook not found: {path}")
        from openpyxl import load_workbook
        org = resolve_organization(options.get("organization_id"))
        wb = load_workbook(path, data_only=True, read_only=False)
        if SHEET_NAME not in wb.sheetnames:
            raise CommandError(f"Sheet '{SHEET_NAME}' missing.")
        rows = rows_by_header(wb[SHEET_NAME])

        rp_lookup = {p.related_party_id: p for p in RelatedParty.objects.filter(organization=org)}
        cp_lookup = {p.counterparty_id: p for p in Counterparty.objects.filter(organization=org)}
        port_lookup = {p.portfolio_id: p for p in Portfolio.objects.filter(organization=org)}
        prop_lookup = {p.property_id: p for p in Property.objects.filter(organization=org)}

        stats = {
            "created": 0, "updated": 0,
            "skipped_no_id": 0, "skipped_explanation_row": 0,
            "linked_owner_rp": 0, "linked_primary_rp": 0,
            "linked_insured_rp": 0, "linked_insurer_cp": 0,
            "linked_portfolio": 0, "linked_property": 0, "linked_beneficiary_rp": 0,
        }
        errors: list[str] = []

        with transaction.atomic():
            for row in rows:
                p_id = text(row.get("Policy_ID"))
                if not p_id:
                    stats["skipped_no_id"] += 1
                    continue
                if is_explanation_row(p_id, prefix="POL_",
                                       name=text(row.get("Policy_Name")),
                                       type_=text(row.get("Policy_Type"))):
                    stats["skipped_explanation_row"] += 1
                    continue

                defaults, metadata = map_row(
                    row, COLUMN_MAP,
                    date_fields=DATE_FIELDS, decimal_fields=DECIMAL_FIELDS,
                    bool_fields=BOOL_FIELDS,
                )
                for k in ("policy_type", "policy_owner_type", "premium_frequency", "status"):
                    if defaults.get(k):
                        defaults[k] = defaults[k].upper()
                clip_short_codes(defaults, currency_fields=("policy_currency", "reporting_currency"))

                def link(code_field, fk_field, lookup, stat_key, master_label):
                    code = text(defaults.get(code_field))
                    if not code:
                        return
                    target = lookup.get(code)
                    if target is not None:
                        defaults[fk_field] = target
                        stats[stat_key] += 1
                    else:
                        errors.append(
                            f"Policy '{p_id}' references unknown {master_label} "
                            f"'{code}' in column {code_field}."
                        )

                link("policy_owner_id_code", "policy_owner_related_party",
                     rp_lookup, "linked_owner_rp", "RelatedParty")
                link("primary_related_party_id_code", "primary_related_party",
                     rp_lookup, "linked_primary_rp", "RelatedParty")
                link("insured_party_id_code", "insured_party_related_party",
                     rp_lookup, "linked_insured_rp", "RelatedParty")
                link("insurer_counterparty_id_code", "insurer_counterparty",
                     cp_lookup, "linked_insurer_cp", "Counterparty")
                link("linked_portfolio_id_code", "linked_portfolio",
                     port_lookup, "linked_portfolio", "Portfolio")
                link("linked_property_id_code", "linked_property",
                     prop_lookup, "linked_property", "Property")
                link("linked_beneficiary_id_code", "linked_beneficiary_related_party",
                     rp_lookup, "linked_beneficiary_rp", "RelatedParty")

                defaults["workbook_metadata"] = metadata
                _, created = Policy.objects.update_or_create(
                    organization=org, policy_id=p_id, defaults=defaults,
                )
                stats["created" if created else "updated"] += 1

            if options["dry_run"]:
                transaction.set_rollback(True)

        prefix = "DRY RUN parsed" if options["dry_run"] else "Imported"
        self.stdout.write(self.style.SUCCESS(f"{prefix} Policy rows for org {org.id} ({org.name})"))
        for k, v in stats.items():
            self.stdout.write(f"  {k}: {v}")
        if errors:
            self.stdout.write(self.style.WARNING(f"{len(errors)} warning(s):"))
            for e in errors[:20]:
                self.stdout.write(f"  - {e}")
