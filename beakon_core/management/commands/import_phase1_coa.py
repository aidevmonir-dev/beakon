from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from beakon_core import constants as c
from beakon_core.models import (
    Account,
    CoADefinition,
    CoAMapping,
    ControlledListEntry,
    Currency,
    DimensionType,
    DimensionValue,
    DimensionValidationRule,
)
from organizations.models import Organization


DEFAULT_WORKBOOK = r"D:\Thomas\2026 04 17-DRAFT-CoA-Wealth management v2.xlsx"


class Command(BaseCommand):
    help = "Import Phase 1 wealth-management CoA tabs from Thomas's workbook."

    def add_arguments(self, parser):
        parser.add_argument(
            "--workbook",
            default=DEFAULT_WORKBOOK,
            help=f"Path to the workbook. Default: {DEFAULT_WORKBOOK}",
        )
        parser.add_argument(
            "--organization-id",
            type=int,
            help="Organization id to import into. If omitted, the only org is used.",
        )
        parser.add_argument(
            "--coa-id",
            default="WM_CLIENT_V1",
            help="CoA_ID to import from the workbook. Default: WM_CLIENT_V1",
        )
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Parse and validate the workbook, then roll back all DB writes.",
        )

    def handle(self, *args, **options):
        workbook_path = Path(options["workbook"])
        if not workbook_path.exists():
            raise CommandError(f"Workbook not found: {workbook_path}")

        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise CommandError(
                "openpyxl is required for this import. Run `pip install -r requirements.txt`."
            ) from exc

        org = _resolve_organization(options.get("organization_id"))
        target_coa_id = options["coa_id"].strip()
        dry_run = options["dry_run"]

        wb = load_workbook(workbook_path, data_only=True, read_only=False)
        required = [
            "01 CoA Definition",
            "02 CoA Master",
            "03 CoA Mapping",
            "04 Dimensions Reference",
            "05 Dimension Values",
            "06 Controlled Lists",
            "09 Dimension Validation Rules",
        ]
        missing = [name for name in required if name not in wb.sheetnames]
        if missing:
            raise CommandError(f"Workbook is missing required Phase 1 tabs: {', '.join(missing)}")

        with transaction.atomic():
            stats = {
                "coa_definitions": 0,
                "accounts": 0,
                "account_parents": 0,
                "dimension_types": 0,
                "dimension_values": 0,
                "controlled_list_entries": 0,
                "coa_mappings": 0,
                "dimension_validation_rules": 0,
                "currencies": 0,
            }

            coa = _import_coa_definition(
                org=org,
                ws=wb["01 CoA Definition"],
                target_coa_id=target_coa_id,
                stats=stats,
            )
            account_lookup = _import_accounts(
                org=org,
                coa=coa,
                ws=wb["02 CoA Master"],
                target_coa_id=target_coa_id,
                stats=stats,
            )
            dimension_lookup = _import_dimension_types(
                org=org,
                ws=wb["04 Dimensions Reference"],
                stats=stats,
            )
            _import_dimension_values(
                org=org,
                ws=wb["05 Dimension Values"],
                dimension_lookup=dimension_lookup,
                stats=stats,
            )
            _import_controlled_lists(
                org=org,
                ws=wb["06 Controlled Lists"],
                stats=stats,
            )
            _import_mappings(
                org=org,
                coa=coa,
                ws=wb["03 CoA Mapping"],
                account_lookup=account_lookup,
                target_coa_id=target_coa_id,
                stats=stats,
            )
            _import_dimension_validation_rules(
                org=org,
                coa=coa,
                ws=wb["09 Dimension Validation Rules"],
                account_lookup=account_lookup,
                target_coa_id=target_coa_id,
                stats=stats,
            )

            for code in _currency_codes_from_definition(coa):
                _, created = Currency.objects.update_or_create(
                    code=code,
                    defaults={
                        "name": _currency_name(code),
                        "symbol": _currency_symbol(code),
                        "decimal_places": 2,
                        "is_active": True,
                    },
                )
                if created:
                    stats["currencies"] += 1

            if dry_run:
                transaction.set_rollback(True)

        prefix = "DRY RUN parsed" if dry_run else "Imported"
        self.stdout.write(self.style.SUCCESS(f"{prefix} Phase 1 CoA for org {org.id} ({org.name})"))
        for key, value in stats.items():
            self.stdout.write(f"  {key}: {value}")


def _resolve_organization(organization_id: int | None) -> Organization:
    if organization_id:
        try:
            return Organization.objects.get(id=organization_id)
        except Organization.DoesNotExist as exc:
            raise CommandError(f"Organization id {organization_id} does not exist.") from exc

    orgs = list(Organization.objects.all()[:2])
    if not orgs:
        raise CommandError("No Organization exists yet. Create one in admin or pass --organization-id.")
    if len(orgs) > 1:
        raise CommandError("Multiple organizations exist. Re-run with --organization-id <id>.")
    return orgs[0]


def _import_coa_definition(*, org, ws, target_coa_id: str, stats: dict[str, int]) -> CoADefinition:
    for row_no in range(2, ws.max_row + 1):
        values = [_value(ws.cell(row_no, col).value) for col in range(1, ws.max_column + 1)]
        if _text(values[0]) != target_coa_id:
            continue

        # Workbook 01 has Additional_Reporting_Currencies split across columns 9
        # and 10 in the only data row (USD, EUR). Universal_Mapping_Required
        # then sits at 11, Dimensions_Enabled at 12, Governed_Instruments_Enabled
        # at 13, Description at 14, free-text notes spilling into 15+.
        additional = ",".join(v for v in (_text(values[9]), _text(values[10])) if v)
        universal_mapping_required = _bool(values[11], default=False)
        dimensions_enabled = _bool(values[12], default=False)
        governed_instruments_enabled = _bool(values[13], default=False)
        description = _text(values[14])
        notes = " ".join(_text(v) for v in values[15:] if _text(v))

        obj, created = CoADefinition.objects.update_or_create(
            organization=org,
            coa_id=target_coa_id,
            defaults={
                "name": _text(values[1]),
                "coa_type": _text(values[2]) or "WEALTH_MGMT",
                "version_no": _int(values[3], default=1),
                "status": _text(values[4]) or "Active",
                "effective_from": _date(values[5]),
                "effective_to": _date(values[6]),
                "base_currency": _text(values[7]) or "CHF",
                "default_reporting_currency": _text(values[8]) or "CHF",
                "additional_reporting_currencies": additional,
                "universal_mapping_required": universal_mapping_required,
                "dimensions_enabled": dimensions_enabled,
                "governed_instruments_enabled": governed_instruments_enabled,
                "description": description,
                "notes": notes,
            },
        )
        stats["coa_definitions"] += 1 if created else 0
        return obj

    raise CommandError(f"CoA_ID {target_coa_id!r} was not found in 01 CoA Definition.")


def _import_accounts(*, org, coa, ws, target_coa_id: str, stats: dict[str, int]) -> dict[str, Account]:
    rows = _rows_by_header(ws)
    lookup: dict[str, Account] = {}
    pending_parents: list[tuple[Account, str]] = []

    for row in rows:
        if _text(row.get("CoA_ID")) != target_coa_id:
            continue
        code = _code(row.get("Account_No"))
        if not code:
            continue

        source_type = _text(row.get("Account_Type")).upper()
        account_type = _map_account_type(source_type, code)
        normal_balance = _normal_balance(row.get("Normal_Balance"), account_type)
        name = _text(row.get("Account_Name"))
        notes = _text(row.get("Notes"))

        account, created = Account.objects.update_or_create(
            organization=org,
            entity=None,
            code=code,
            defaults={
                "coa_definition": coa,
                "name": name,
                "short_name": _text(row.get("Short_Name")),
                "source_account_type": source_type,
                "account_type": account_type,
                "account_subtype": _infer_subtype(code, source_type, name),
                "normal_balance": normal_balance,
                "currency": "",
                "level_no": _int(row.get("Level_No"), default=None),
                "posting_allowed": _bool(row.get("Posting_Allowed"), default=True),
                "header_flag": _bool(row.get("Header_Flag"), default=False),
                "mandatory_flag": _bool(row.get("Mandatory_Flag"), default=False),
                "scope": _text(row.get("Scope")),
                "report_group": _text(row.get("Report_Group")),
                "cashflow_category": _text(row.get("Cashflow_Category")),
                "required_dimension_type_codes": _text(row.get("Required_Dimension_Type_Codes")),
                "optional_dimension_type_codes": _text(row.get("Optional_Dimension_Type_Codes")),
                "dimension_validation_rule": _text(row.get("Dimension_Validation_Rule")),
                "universal_map_required": _bool(row.get("Universal_Map_Required"), default=False),
                "default_universal_coa_code": _text(row.get("Default_Universal_CoA_Code")),
                "mapping_method": _text(row.get("Mapping_Method")),
                "fx_revalue_flag": _bool(row.get("FX_Revalue_Flag"), default=False),
                "tax_relevant_flag": _bool(row.get("Tax_Relevant_Flag"), default=False),
                "monetary_flag": _bool(row.get("Monetary_Flag"), default=False),
                # Governance fields promoted from JSON. Each value is filtered
                # against a controlled vocabulary; off-vocab values fall through
                # as blank but the raw value remains in workbook_metadata.
                "economic_nature": _vocab_filtered(row.get("Economic_Nature"), _ECONOMIC_NATURES),
                "client_view_category": _short_text(row.get("Client_View_Category"), max_len=60),
                "subledger_required": _bool(row.get("Subledger_Required"), default=False),
                "subledger_type": _vocab_filtered(row.get("Subledger_Type"), _SUBLEDGER_TYPES),
                "valuation_basis": _vocab_filtered(row.get("Valuation_Basis"), _VALUATION_BASES),
                "allow_foreign_currency_posting": _bool(
                    row.get("Allow_Foreign_Currency_Posting"), default=True),
                "fx_remeasure_flag": _vocab_filtered(
                    row.get("FX_Remeasure_Flag"), _FX_REMEASURE_FLAGS),
                "fx_gain_loss_bucket": _vocab_filtered(
                    row.get("FX_Gain_Loss_Bucket"), _FX_GAIN_LOSS_BUCKETS),
                "historical_vs_closing_rate_method": _vocab_filtered(
                    row.get("Historical_vs_Closing_Rate_Method"), _HIST_RATE_METHODS),
                "is_active": _bool(row.get("Active_Flag"), default=True),
                "is_system": False,
                "description": notes,
                "workbook_metadata": _json_safe(row),
            },
        )
        lookup[code] = account
        parent_code = _code(row.get("Parent_Account_No"))
        if parent_code:
            pending_parents.append((account, parent_code))
        if created:
            stats["accounts"] += 1

    for account, parent_code in pending_parents:
        parent = lookup.get(parent_code)
        if parent and account.parent_id != parent.id:
            account.parent = parent
            account.save(update_fields=["parent", "updated_at"])
            stats["account_parents"] += 1

    return lookup


def _import_dimension_types(*, org, ws, stats: dict[str, int]) -> dict[str, DimensionType]:
    lookup: dict[str, DimensionType] = {}
    for row_no in range(2, ws.max_row + 1):
        values = [_value(ws.cell(row_no, col).value) for col in range(1, ws.max_column + 1)]
        code = _text(values[0])
        if not code:
            continue

        tail = [_text(v) for v in values[3:] if _text(v)]
        applies_to = ""
        mandatory = multi = hierarchy = active = False
        owner = notes = ""
        if len(tail) >= 6:
            applies_to = ",".join(tail[:-6])
            mandatory = _bool(tail[-6], default=False)
            multi = _bool(tail[-5], default=False)
            owner = tail[-4]
            hierarchy = _bool(tail[-3], default=False)
            active = _bool(tail[-2], default=True)
            notes = tail[-1]

        obj, created = DimensionType.objects.update_or_create(
            organization=org,
            code=code,
            defaults={
                "name": _text(values[1]),
                "description": _text(values[2]),
                "applies_to": applies_to,
                "mandatory_flag": mandatory,
                "multi_select_allowed": multi,
                "master_data_owner": owner,
                "hierarchy_allowed": hierarchy,
                "active_flag": active,
                "notes": notes,
                "workbook_metadata": {"row": row_no, "values": _json_safe(values)},
            },
        )
        lookup[code] = obj
        if created:
            stats["dimension_types"] += 1
    return lookup


def _import_dimension_values(*, org, ws, dimension_lookup, stats: dict[str, int]) -> None:
    for row in _rows_by_header(ws):
        type_code = _text(row.get("Dimension_Type_Code"))
        value_code = _text(row.get("Dimension_Value_Code"))
        if not type_code or not value_code:
            continue
        dimension_type = dimension_lookup.get(type_code)
        if dimension_type is None:
            dimension_type, _ = DimensionType.objects.update_or_create(
                organization=org,
                code=type_code,
                defaults={"name": type_code, "active_flag": True},
            )
            dimension_lookup[type_code] = dimension_type

        _, created = DimensionValue.objects.update_or_create(
            organization=org,
            dimension_type=dimension_type,
            code=value_code,
            defaults={
                "name": _text(row.get("Dimension_Value_Name")),
                "parent_value_code": _text(row.get("Parent_Value_Code")),
                "description": _text(row.get("Description")),
                "active_flag": _bool(row.get("Active_Flag"), default=True),
                "effective_from": _date(row.get("Effective_From")),
                "effective_to": _date(row.get("Effective_To")),
                "external_reference": _text(row.get("External_Reference")),
                "notes": _text(row.get("Notes")),
                "workbook_metadata": _json_safe(row),
            },
        )
        if created:
            stats["dimension_values"] += 1


def _import_controlled_lists(*, org, ws, stats: dict[str, int]) -> None:
    for row in _rows_by_header(ws):
        list_name = _text(row.get("List_Name"))
        list_code = _text(row.get("List_Code"))
        if not list_name or not list_code:
            continue
        _, created = ControlledListEntry.objects.update_or_create(
            organization=org,
            list_name=list_name,
            list_code=list_code,
            defaults={
                "list_value": _text(row.get("List_Value")),
                "display_order": _int(row.get("Display_Order"), default=0) or 0,
                "active_flag": _bool(row.get("Active_Flag"), default=True),
                "description": _text(row.get("Description")),
                "notes": _text(row.get("Notes")),
            },
        )
        if created:
            stats["controlled_list_entries"] += 1


def _import_mappings(*, org, coa, ws, account_lookup, target_coa_id: str, stats: dict[str, int]) -> None:
    for row in _rows_by_header(ws):
        if _text(row.get("Source_CoA_ID")) != target_coa_id:
            continue
        mapping_id = _text(row.get("Mapping_ID"))
        source_no = _code(row.get("Source_Account_No"))
        universal_code = _text(row.get("Universal_CoA_Code"))
        if not mapping_id or not source_no or not universal_code:
            continue

        _, created = CoAMapping.objects.update_or_create(
            organization=org,
            mapping_id=mapping_id,
            defaults={
                "coa_definition": coa,
                "account": account_lookup.get(source_no),
                "source_account_no": source_no,
                "source_account_name": _text(row.get("Source_Account_Name")),
                "universal_coa_code": universal_code,
                "universal_coa_name": _text(row.get("Universal_CoA_Name")),
                "mapping_type": _text(row.get("Mapping_Type")),
                "mapping_percent": _decimal(row.get("Mapping_Percent"), default=Decimal("100")),
                "condition_rule": _text(row.get("Condition_Rule")),
                "required_dimension": _text(row.get("Required_Dimension")),
                "effective_from": _date(row.get("Effective_From")),
                "effective_to": _date(row.get("Effective_To")),
                "review_status": _text(row.get("Review_Status")),
                "approved_by": _text(row.get("Approved_By")),
                "notes": _text(row.get("Notes")),
                "workbook_metadata": _json_safe(row),
            },
        )
        if created:
            stats["coa_mappings"] += 1


def _import_dimension_validation_rules(*, org, coa, ws, account_lookup, target_coa_id: str, stats: dict[str, int]) -> None:
    for row in _rows_by_header(ws):
        if _text(row.get("CoA_ID")) != target_coa_id:
            continue
        rule_id = _text(row.get("Rule_ID"))
        account_no = _code(row.get("Account_No"))
        if not rule_id or not account_no:
            continue

        _, created = DimensionValidationRule.objects.update_or_create(
            organization=org,
            rule_id=rule_id,
            defaults={
                "coa_definition": coa,
                "account": account_lookup.get(account_no),
                "account_no": account_no,
                "account_name": _text(row.get("Account_Name")),
                "rule_type": _text(row.get("Rule_Type")),
                "trigger_event": _text(row.get("Trigger_Event")),
                "required_dimension_type_codes": _text(row.get("Required_Dimension_Type_Codes")),
                "optional_dimension_type_codes": _text(row.get("Optional_Dimension_Type_Codes")),
                "conditional_dimension_type_codes": _text(row.get("Conditional_Dimension_Type_Codes")),
                "condition_expression": _text(row.get("Condition_Expression")),
                "validation_error_message": _text(row.get("Validation_Error_Message")),
                "severity": _text(row.get("Severity")),
                "master_driver": _text(row.get("Master_Driver")),
                "active_flag": _bool(row.get("Active_Flag"), default=True),
                "effective_from": _date(row.get("Effective_From")),
                "effective_to": _date(row.get("Effective_To")),
                "notes": _text(row.get("Notes")),
                "workbook_metadata": _json_safe(row),
            },
        )
        if created:
            stats["dimension_validation_rules"] += 1


def _rows_by_header(ws) -> list[dict[str, Any]]:
    headers = [_text(ws.cell(1, col).value) for col in range(1, ws.max_column + 1)]
    rows = []
    for row_no in range(2, ws.max_row + 1):
        row: dict[str, Any] = {"_row": row_no}
        has_value = False
        for idx, header in enumerate(headers, start=1):
            value = _value(ws.cell(row_no, idx).value)
            if value not in ("", None):
                has_value = True
            if header:
                row[header] = value
        if has_value:
            rows.append(row)
    return rows


def _currency_codes_from_definition(coa: CoADefinition) -> set[str]:
    codes = {coa.base_currency, coa.default_reporting_currency}
    codes.update(part.strip() for part in coa.additional_reporting_currencies.split(","))
    return {code.upper() for code in codes if code}


def _currency_name(code: str) -> str:
    return {
        "CHF": "Swiss Franc",
        "USD": "US Dollar",
        "EUR": "Euro",
        "GBP": "Pound Sterling",
    }.get(code, code)


def _currency_symbol(code: str) -> str:
    return {
        "CHF": "CHF",
        "USD": "$",
        "EUR": "EUR",
        "GBP": "GBP",
    }.get(code, code)


def _map_account_type(source_type: str, code: str) -> str:
    source_type = source_type.upper()
    if source_type in {"ASSET", "CONTRA_ASSET"}:
        return c.ACCOUNT_TYPE_ASSET
    if source_type == "LIABILITY":
        return c.ACCOUNT_TYPE_LIABILITY
    if source_type == "EQUITY":
        return c.ACCOUNT_TYPE_EQUITY
    if source_type == "INCOME":
        return c.ACCOUNT_TYPE_REVENUE
    if source_type == "EXPENSE":
        return c.ACCOUNT_TYPE_EXPENSE

    first = code[:1]
    if first == "1":
        return c.ACCOUNT_TYPE_ASSET
    if first == "2":
        return c.ACCOUNT_TYPE_LIABILITY
    if first == "3":
        return c.ACCOUNT_TYPE_EQUITY
    if first == "4":
        return c.ACCOUNT_TYPE_REVENUE
    return c.ACCOUNT_TYPE_EXPENSE


def _normal_balance(raw: Any, account_type: str) -> str:
    value = _text(raw).lower()
    if value in {c.NORMAL_BALANCE_DEBIT, c.NORMAL_BALANCE_CREDIT}:
        return value
    return c.NORMAL_BALANCE_MAP.get(account_type, c.NORMAL_BALANCE_DEBIT)


def _infer_subtype(code: str, source_type: str, name: str) -> str:
    text = f"{code} {source_type} {name}".lower()
    if source_type == "HEADER":
        return ""
    if "accounts payable" in text or "payable" in text:
        return "accounts_payable"
    if "accounts receivable" in text or "receivable" in text:
        return "accounts_receivable"
    if "due from" in text:
        return "intercompany_receivable"
    if "due to" in text:
        return "intercompany_payable"
    if "loan" in text or "mortgage" in text or "facility" in text:
        return "loan_payable" if code.startswith("2") else "loan_receivable"
    if "bank" in text or code.startswith("1102") or code.startswith("1103"):
        return "bank"
    if "cash" in text or code.startswith("1101"):
        return "cash"
    if source_type in {"ASSET", "CONTRA_ASSET"}:
        if code.startswith(("13", "14", "15")) or "investment" in text:
            return "investment"
        return "other_asset"
    if source_type == "LIABILITY":
        return "other_liability"
    if source_type == "EQUITY":
        return "other_equity"
    if source_type == "INCOME":
        return "investment_income" if "dividend" in text or "interest" in text else "other_income"
    if source_type == "EXPENSE":
        return "tax_expense" if "tax" in text else "other_expense"
    return ""


def _value(value: Any) -> Any:
    if isinstance(value, str):
        return value.strip()
    return value


def _text(value: Any) -> str:
    value = _value(value)
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, Decimal):
        return str(value)
    return str(value).strip()


def _code(value: Any) -> str:
    text = _text(value)
    if text.endswith(".0"):
        text = text[:-2]
    return text


def _bool(value: Any, *, default: bool) -> bool:
    text = _text(value).lower()
    if not text:
        return default
    if text in {"yes", "y", "true", "1", "active"}:
        return True
    if text in {"no", "n", "false", "0", "inactive"}:
        return False
    return default


def _int(value: Any, *, default: int | None) -> int | None:
    if value in ("", None):
        return default
    try:
        return int(value)
    except (TypeError, ValueError):
        return default


def _decimal(value: Any, *, default: Decimal) -> Decimal:
    if value in ("", None):
        return default
    try:
        return Decimal(str(value))
    except Exception:
        return default


def _date(value: Any) -> date | None:
    if value in ("", None):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    try:
        return datetime.strptime(_text(value), "%Y-%m-%d").date()
    except ValueError:
        return None


def _json_safe(value: Any) -> Any:
    if isinstance(value, dict):
        return {str(k): _json_safe(v) for k, v in value.items() if not str(k).startswith("_")}
    if isinstance(value, list):
        return [_json_safe(v) for v in value]
    if isinstance(value, tuple):
        return [_json_safe(v) for v in value]
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, Decimal):
        return str(value)
    return value


# ── Controlled vocabularies for Account governance fields ────────────────

_ECONOMIC_NATURES = {
    "BALANCE_SHEET", "ACCRUAL", "PORTFOLIO_ASSET", "RELATED_PARTY_BALANCE",
    "SETTLEMENT", "IMPAIRMENT", "CAPITAL_MOVEMENT", "CLEARING", "TAX",
    "PREPAYMENT", "VALUATION_MOVEMENT", "PERFORMANCE",
}
_SUBLEDGER_TYPES = {
    "NONE", "BANK", "INVESTMENT", "RELATED_PARTY", "COUNTERPARTY",
    "PROPERTY", "FIXED_ASSET",
}
_VALUATION_BASES = {
    "NA", "FVPL", "FAIR_VALUE", "AMORTIZED_COST", "COST", "HISTORICAL",
}
_FX_REMEASURE_FLAGS = {"YES", "NO", "BY_VALUATION_METHOD"}
_FX_GAIN_LOSS_BUCKETS = {
    "NA", "UNREALIZED_FX", "COMBINED_REALIZED_UNREALIZED", "FAIR_VALUE_FX",
}
_HIST_RATE_METHODS = {
    "NA", "CLOSING", "AVERAGE", "HISTORICAL", "FAIR_VALUE",
    "BY_VALUATION_METHOD",
}


def _vocab_filtered(value: Any, allowed: set[str]) -> str:
    """Return the value only if it sits in the controlled vocabulary."""
    if value is None:
        return ""
    s = str(value).strip().upper()
    return s if s in allowed else ""


def _short_text(value: Any, *, max_len: int) -> str:
    """Accept a short controlled-list-style value, reject free-text noise."""
    if value is None:
        return ""
    s = str(value).strip()
    if not s or len(s) > max_len:
        return ""
    if " " in s and "_" not in s:
        return ""
    if not all(c.isalnum() or c == "_" for c in s):
        return ""
    return s.upper()
