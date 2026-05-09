"""Import workbook tab `10_Counterparty_Master`."""
from __future__ import annotations

from pathlib import Path
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from beakon_core.models import Counterparty
from beakon_core.management.commands._master_import_helpers import (
    is_explanation_row, map_row, resolve_organization, rows_by_header, text,
)

DEFAULT_WORKBOOK = r"D:\Thomas\2026 04 17-DRAFT-CoA-Wealth management v2.xlsx"
SHEET_NAME = "10_Counterparty_Master"

DATE_FIELDS = set()
DECIMAL_FIELDS = set()
BOOL_FIELDS = {
    "active_flag", "sanctions_check_flag", "related_party_flag",
    "intercompany_flag", "loan_eligible_flag", "ap_eligible_flag",
    "ar_eligible_flag", "tax_eligible_flag", "insurance_eligible_flag",
    "education_eligible_flag", "professional_fees_eligible_flag",
}

COLUMN_MAP = {
    "Counterparty_ID": "counterparty_id",
    "Counterparty_Name": "counterparty_name",
    "Short_Name": "short_name",
    "Counterparty_Type": "counterparty_type",
    "Counterparty_Subtype": "counterparty_subtype",
    "Status": "status",
    "Active_Flag": "active_flag",
    "External_Reference": "external_reference",
    "Tax_ID": "tax_id",
    "Registration_No": "registration_no",
    "Country_Code": "country_code",
    "Jurisdiction_Code": "jurisdiction_code",
    "Base_Currency": "base_currency",
    "Default_Payment_Currency": "default_payment_currency",
    "Language_Code": "language_code",
    "Payment_Terms": "payment_terms",
    "Settlement_Method": "settlement_method",
    "Default_Bank_Reference": "default_bank_reference",
    "Risk_Rating": "risk_rating",
    "KYC_Status": "kyc_status",
    "AML_Risk_Level": "aml_risk_level",
    "Sanctions_Check_Flag": "sanctions_check_flag",
    "Related_Party_Flag": "related_party_flag",
    "Intercompany_Flag": "intercompany_flag",
    "Loan_Eligible_Flag": "loan_eligible_flag",
    "AP_Eligible_Flag": "ap_eligible_flag",
    "AR_Eligible_Flag": "ar_eligible_flag",
    "Tax_Eligible_Flag": "tax_eligible_flag",
    "Insurance_Eligible_Flag": "insurance_eligible_flag",
    "Education_Eligible_Flag": "education_eligible_flag",
    "Professional_Fees_Eligible_Flag": "professional_fees_eligible_flag",
    "Notes": "notes",
}


class Command(BaseCommand):
    help = "Import tab 10_Counterparty_Master."

    def add_arguments(self, parser):
        parser.add_argument("--workbook", default=DEFAULT_WORKBOOK)
        parser.add_argument("--organization-id", type=int)
        parser.add_argument("--dry-run", action="store_true")

    def handle(self, *args, **options):
        path = Path(options["workbook"])
        if not path.exists():
            raise CommandError(f"Workbook not found: {path}")
        from openpyxl import load_workbook
        org = resolve_organization(options.get("organization_id"))
        wb = load_workbook(path, data_only=True, read_only=False)
        if SHEET_NAME not in wb.sheetnames:
            raise CommandError(f"Sheet '{SHEET_NAME}' missing.")
        rows = rows_by_header(wb[SHEET_NAME])

        stats = {"created": 0, "updated": 0, "skipped_no_id": 0, "skipped_explanation_row": 0}

        with transaction.atomic():
            for row in rows:
                cp_id = text(row.get("Counterparty_ID"))
                if not cp_id:
                    stats["skipped_no_id"] += 1
                    continue
                if is_explanation_row(cp_id, prefix="CP_",
                                       name=text(row.get("Counterparty_Name")),
                                       type_=text(row.get("Counterparty_Type"))):
                    stats["skipped_explanation_row"] += 1
                    continue

                defaults, metadata = map_row(
                    row, COLUMN_MAP,
                    date_fields=DATE_FIELDS, decimal_fields=DECIMAL_FIELDS,
                    bool_fields=BOOL_FIELDS,
                )
                for k in ("counterparty_type", "status"):
                    if defaults.get(k):
                        defaults[k] = defaults[k].upper()
                if defaults.get("base_currency"):
                    defaults["base_currency"] = defaults["base_currency"].upper()[:3]
                if defaults.get("default_payment_currency"):
                    defaults["default_payment_currency"] = defaults["default_payment_currency"].upper()[:3]
                defaults["workbook_metadata"] = metadata

                _, created = Counterparty.objects.update_or_create(
                    organization=org, counterparty_id=cp_id, defaults=defaults,
                )
                stats["created" if created else "updated"] += 1

            if options["dry_run"]:
                transaction.set_rollback(True)

        prefix = "DRY RUN parsed" if options["dry_run"] else "Imported"
        self.stdout.write(self.style.SUCCESS(f"{prefix} Counterparty rows for org {org.id} ({org.name})"))
        for k, v in stats.items():
            self.stdout.write(f"  {k}: {v}")
