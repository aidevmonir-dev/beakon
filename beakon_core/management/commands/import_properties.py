"""Import workbook tab `15_Property_Master`."""
from __future__ import annotations

from pathlib import Path
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from beakon_core.models import Loan, Portfolio, Property, RelatedParty
from beakon_core.management.commands._master_import_helpers import (
    clip_short_codes, is_explanation_row, map_row, resolve_organization,
    rows_by_header, text,
)

DEFAULT_WORKBOOK = r"D:\Thomas\2026 04 17-DRAFT-CoA-Wealth management v2.xlsx"
SHEET_NAME = "15_Property_Master"

DATE_FIELDS = {"acquisition_date", "disposal_date", "valuation_date"}
DECIMAL_FIELDS = {"acquisition_cost", "current_carrying_value"}
BOOL_FIELDS = {
    "active_flag", "posting_allowed_flag",
    "approval_required_flag", "source_document_required_flag",
    "mortgage_linked_flag", "rental_income_flag", "personal_use_flag",
    "expense_allocation_allowed_flag", "net_worth_inclusion_flag",
}

COLUMN_MAP = {
    "Property_ID": "property_id",
    "Property_Name": "property_name",
    "Short_Name": "short_name",
    "Property_Type": "property_type",
    "Property_Subtype": "property_subtype",
    "Usage_Type": "usage_type",
    "Ownership_Type": "ownership_type",
    "Owner_Type": "owner_type",
    "Owner_ID": "owner_id_code",
    "Primary_Related_Party_ID": "primary_related_party_id_code",
    "Linked_Portfolio_ID": "linked_portfolio_id_code",
    "Linked_SPV_ID": "linked_spv_id",
    "Address_Line_1": "address_line_1",
    "Address_Line_2": "address_line_2",
    "City": "city",
    "Region": "region",
    "Postal_Code": "postal_code",
    "Country_Code": "country_code",
    "Jurisdiction_Code": "jurisdiction_code",
    "Property_Currency": "property_currency",
    "Reporting_Currency": "reporting_currency",
    "Acquisition_Date": "acquisition_date",
    "Disposal_Date": "disposal_date",
    "Acquisition_Cost": "acquisition_cost",
    "Current_Carrying_Value": "current_carrying_value",
    "Valuation_Method": "valuation_method",
    "Valuation_Date": "valuation_date",
    "Mortgage_Linked_Flag": "mortgage_linked_flag",
    "Linked_Loan_ID": "linked_loan_id_code",
    "Rental_Income_Flag": "rental_income_flag",
    "Personal_Use_Flag": "personal_use_flag",
    "Expense_Allocation_Allowed_Flag": "expense_allocation_allowed_flag",
    "Net_Worth_Inclusion_Flag": "net_worth_inclusion_flag",
    "Status": "status",
    "Active_Flag": "active_flag",
    "Posting_Allowed_Flag": "posting_allowed_flag",
    "Approval_Required_Flag": "approval_required_flag",
    "Source_Document_Required_Flag": "source_document_required_flag",
    "Notes": "notes",
}


class Command(BaseCommand):
    help = "Import tab 15_Property_Master."

    def add_arguments(self, parser):
        parser.add_argument("--workbook", default=DEFAULT_WORKBOOK)
        parser.add_argument("--organization-id", type=int)
        parser.add_argument("--dry-run", action="store_true")

    def handle(self, *args, **options):
        path = Path(options["workbook"])
        if not path.exists():
            raise CommandError(f"Workbook not found: {path}")
        from openpyxl import load_workbook
        org = resolve_organization(options.get("organization_id"))
        wb = load_workbook(path, data_only=True, read_only=False)
        if SHEET_NAME not in wb.sheetnames:
            raise CommandError(f"Sheet '{SHEET_NAME}' missing.")
        rows = rows_by_header(wb[SHEET_NAME])

        rp_lookup = {p.related_party_id: p for p in RelatedParty.objects.filter(organization=org)}
        port_lookup = {p.portfolio_id: p for p in Portfolio.objects.filter(organization=org)}
        loan_lookup = {p.loan_id: p for p in Loan.objects.filter(organization=org)}

        stats = {
            "created": 0, "updated": 0,
            "skipped_no_id": 0, "skipped_explanation_row": 0,
            "linked_owner_rp": 0, "linked_primary_rp": 0,
            "linked_portfolio": 0, "linked_loan": 0,
        }
        errors: list[str] = []

        with transaction.atomic():
            for row in rows:
                p_id = text(row.get("Property_ID"))
                if not p_id:
                    stats["skipped_no_id"] += 1
                    continue
                if is_explanation_row(p_id, prefix="PROP_",
                                       name=text(row.get("Property_Name")),
                                       type_=text(row.get("Property_Type"))):
                    stats["skipped_explanation_row"] += 1
                    continue

                defaults, metadata = map_row(
                    row, COLUMN_MAP,
                    date_fields=DATE_FIELDS, decimal_fields=DECIMAL_FIELDS,
                    bool_fields=BOOL_FIELDS,
                )
                for k in ("property_type", "usage_type", "ownership_type", "owner_type", "status"):
                    if defaults.get(k):
                        defaults[k] = defaults[k].upper()
                clip_short_codes(defaults, currency_fields=("property_currency", "reporting_currency"))

                def link(code_field, fk_field, lookup, stat_key, master_label):
                    code = text(defaults.get(code_field))
                    if not code:
                        return
                    target = lookup.get(code)
                    if target is not None:
                        defaults[fk_field] = target
                        stats[stat_key] += 1
                    else:
                        errors.append(
                            f"Property '{p_id}' references unknown {master_label} "
                            f"'{code}' in column {code_field}."
                        )

                link("owner_id_code", "owner_related_party",
                     rp_lookup, "linked_owner_rp", "RelatedParty")
                link("primary_related_party_id_code", "primary_related_party",
                     rp_lookup, "linked_primary_rp", "RelatedParty")
                link("linked_portfolio_id_code", "linked_portfolio",
                     port_lookup, "linked_portfolio", "Portfolio")
                link("linked_loan_id_code", "linked_loan",
                     loan_lookup, "linked_loan", "Loan")

                defaults["workbook_metadata"] = metadata
                _, created = Property.objects.update_or_create(
                    organization=org, property_id=p_id, defaults=defaults,
                )
                stats["created" if created else "updated"] += 1

            if options["dry_run"]:
                transaction.set_rollback(True)

        prefix = "DRY RUN parsed" if options["dry_run"] else "Imported"
        self.stdout.write(self.style.SUCCESS(f"{prefix} Property rows for org {org.id} ({org.name})"))
        for k, v in stats.items():
            self.stdout.write(f"  {k}: {v}")
        if errors:
            self.stdout.write(self.style.WARNING(f"{len(errors)} warning(s):"))
            for e in errors[:20]:
                self.stdout.write(f"  - {e}")
