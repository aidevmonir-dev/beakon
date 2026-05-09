"""AICoAImportService — turn an uploaded CoA file (xlsx/csv) into structured
Account rows the user can review + commit.

Mirrors the AIBillDraftingService pattern: AI proposes (with reasoning per
row), user reviews, the kernel writes. Nothing is committed without
explicit user action.

Workflow:
    preview(file_bytes, content_type, filename)
        1. Parse the file into raw text rows (sheet -> rows of cells).
        2. Send the raw rows + the valid account_type / subtype menu to
           Claude via tool_use.
        3. Claude returns one structured row per CoA entry, with
           confidence + a short rationale (so the reviewer learns why).
        4. Service returns the array unchanged for the UI to render.

    commit(organization, entity, rows, user)
        1. Validate every row against ACCOUNT_TYPE_CHOICES /
           ACCOUNT_SUBTYPE_CHOICES, dedupe vs. existing accounts on the
           same (org, entity, code).
        2. bulk_create Account rows in one transaction.
        3. Log an AuditEvent with the full proposal -> commit diff.

The AI is never the source of truth. Every committed row is the user's
explicit choice (the UI lets them edit any cell before commit).
"""
from __future__ import annotations

import csv
import io
import json
from decimal import Decimal
from typing import Any, Optional

from django.conf import settings
from django.db import transaction

from audit.services import log_event

from .. import constants as c
from ..exceptions import ValidationError
from ..models import Account, Entity


VALID_ACCOUNT_TYPES = [v for v, _ in c.ACCOUNT_TYPE_CHOICES]
VALID_SUBTYPES = [v for v, _ in c.ACCOUNT_SUBTYPE_CHOICES]
SUBTYPES_BY_TYPE = {
    c.ACCOUNT_TYPE_ASSET: [
        "bank", "cash", "current_asset", "accounts_receivable",
        "intercompany_receivable", "prepaid", "inventory", "investment",
        "loan_receivable", "vat_receivable", "tax_receivable",
        "fixed_asset", "accumulated_depreciation", "intangible_asset",
        "other_asset",
    ],
    c.ACCOUNT_TYPE_LIABILITY: [
        "accounts_payable", "intercompany_payable", "accrued_liability",
        "current_liability", "loan_payable", "long_term_liability",
        "tax_payable", "vat_payable", "other_liability",
    ],
    c.ACCOUNT_TYPE_EQUITY: [
        "capital", "retained_earnings", "revaluation_reserve",
        "fx_translation_reserve", "distribution", "other_equity",
    ],
    c.ACCOUNT_TYPE_REVENUE: [
        "operating_revenue", "investment_income", "fx_gain", "other_income",
    ],
    c.ACCOUNT_TYPE_EXPENSE: [
        "cogs", "operating_expense", "professional_fees", "depreciation",
        "fx_loss", "tax_expense", "other_expense",
    ],
}

# Cap how many rows we send to Claude in one call. Family-office workbooks
# can exceed 1000 rows; chunking keeps responses fast and within token
# limits. The UI calls preview() once per chunk and concatenates.
MAX_ROWS_PER_CALL = 200

# Anthropic tool schema. Strict JSON shape so the model can't drift.
_COA_TOOL = {
    "name": "extract_coa",
    "description": (
        "Return a structured Chart of Accounts derived from the uploaded "
        "file. Each row must have a unique code within the entity. "
        "Use the closest matching account_type and account_subtype from "
        "the allowed lists. If a row in the source is a HEADER (no "
        "postings, just a section title), set is_header=true and leave "
        "subtype empty."
    ),
    "input_schema": {
        "type": "object",
        "required": ["accounts"],
        "properties": {
            "accounts": {
                "type": "array",
                "items": {
                    "type": "object",
                    "required": ["code", "name", "account_type", "confidence"],
                    "properties": {
                        "code": {"type": "string"},
                        "name": {"type": "string"},
                        "account_type": {
                            "type": "string",
                            "enum": VALID_ACCOUNT_TYPES,
                        },
                        "account_subtype": {
                            "type": "string",
                            "enum": VALID_SUBTYPES + [""],
                        },
                        "parent_code": {"type": ["string", "null"]},
                        "is_header": {"type": "boolean"},
                        "description": {"type": "string"},
                        "confidence": {
                            "type": "number",
                            "minimum": 0,
                            "maximum": 1,
                            "description": (
                                "0.0-1.0 — how sure the model is about "
                                "the type/subtype classification."
                            ),
                        },
                        "rationale": {
                            "type": "string",
                            "description": (
                                "One short sentence explaining why this "
                                "type/subtype was picked."
                            ),
                        },
                    },
                },
            },
        },
    },
}


class AICoAImportService:
    @staticmethod
    def preview(
        *, file_bytes: bytes, content_type: str, filename: str,
    ) -> dict:
        """Parse + classify. Returns ``{accounts: [...], source: {...}}``."""
        rows, source_meta = _parse_file(file_bytes, content_type, filename)
        if not rows:
            raise ValidationError(
                "Could not find any rows in the uploaded file.",
                code="COA001",
            )
        # Use the first row as headers if it looks like text. Many
        # workbooks have a title row above the table — leave that to the
        # AI to figure out.
        chunks = [rows[i:i + MAX_ROWS_PER_CALL] for i in range(0, len(rows), MAX_ROWS_PER_CALL)]
        all_accounts: list[dict] = []
        for chunk in chunks:
            extracted = _ask_claude(chunk)
            all_accounts.extend(extracted)
        return {
            "accounts": all_accounts,
            "source": source_meta,
        }

    @staticmethod
    @transaction.atomic
    def commit(
        *,
        organization,
        entity: Optional[Entity],
        rows: list[dict],
        user=None,
    ) -> dict:
        """Write reviewed rows as Account records.

        ``entity=None`` -> shared accounts (Account.entity is NULL).
        Skips rows where ``is_header`` is true (those are display-only).
        Skips rows whose (org, entity, code) already exists — never
        overwrites.
        """
        existing_codes = set(
            Account.objects
            .filter(organization=organization, entity=entity)
            .values_list("code", flat=True)
        )
        to_create: list[Account] = []
        skipped: list[dict] = []
        errors: list[dict] = []
        currency = (entity.functional_currency if entity else "") or ""

        for row in rows:
            try:
                if row.get("is_header"):
                    skipped.append({"code": row.get("code"), "reason": "header"})
                    continue
                code = (row.get("code") or "").strip()
                name = (row.get("name") or "").strip()
                atype = (row.get("account_type") or "").strip()
                sub = (row.get("account_subtype") or "").strip()
                if not code or not name or not atype:
                    errors.append({"code": code, "reason": "missing required field"})
                    continue
                if atype not in VALID_ACCOUNT_TYPES:
                    errors.append({"code": code, "reason": f"invalid account_type {atype!r}"})
                    continue
                if sub and sub not in VALID_SUBTYPES:
                    errors.append({"code": code, "reason": f"invalid account_subtype {sub!r}"})
                    continue
                if code in existing_codes:
                    skipped.append({"code": code, "reason": "already exists"})
                    continue
                to_create.append(Account(
                    organization=organization,
                    entity=entity,
                    code=code,
                    name=name,
                    account_type=atype,
                    account_subtype=sub,
                    normal_balance=c.NORMAL_BALANCE_MAP.get(
                        atype, c.NORMAL_BALANCE_DEBIT,
                    ),
                    currency=currency,
                    is_active=True,
                    posting_allowed=not row.get("is_header", False),
                    description=row.get("description") or "",
                ))
                existing_codes.add(code)  # avoid intra-batch duplicates
            except Exception as e:
                errors.append({"code": row.get("code"), "reason": str(e)})

        if to_create:
            Account.objects.bulk_create(to_create)

        log_event(
            organization=organization,
            actor=user,
            actor_type="ai",
            action="coa_import.commit",
            object_type="Account",
            object_id=entity.id if entity else organization.id,
            object_repr=f"CoA import -> {entity.code if entity else 'shared'} ({len(to_create)} accounts)",
            metadata={
                "entity_id": entity.id if entity else None,
                "entity_code": entity.code if entity else None,
                "rows_total": len(rows),
                "created": len(to_create),
                "skipped": skipped,
                "errors": errors,
            },
        )
        return {
            "created": len(to_create),
            "skipped": skipped,
            "errors": errors,
        }


# ── Helpers ──────────────────────────────────────────────────────────────

def _parse_file(file_bytes: bytes, content_type: str, filename: str) -> tuple[list[list[Any]], dict]:
    """Return (rows, source_meta). Each row is a list of stringified cells."""
    name_lower = (filename or "").lower()
    is_xlsx = (
        "spreadsheetml" in (content_type or "")
        or name_lower.endswith(".xlsx")
        or name_lower.endswith(".xlsm")
    )
    is_csv = (
        "csv" in (content_type or "")
        or name_lower.endswith(".csv")
        or name_lower.endswith(".tsv")
    )
    if is_xlsx:
        return _parse_xlsx(file_bytes, filename)
    if is_csv:
        return _parse_csv(file_bytes, filename)
    raise ValidationError(
        f"Unsupported file type: {content_type or filename}. "
        "Upload .xlsx or .csv.",
        code="COA002",
    )


def _parse_xlsx(file_bytes: bytes, filename: str) -> tuple[list[list[Any]], dict]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ValidationError(
            "openpyxl is required to parse .xlsx files.",
            code="COA003",
        ) from exc
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    # Pick the sheet with the most non-empty rows — heuristic that beats
    # asking the user to pick a sheet name. Skips obviously empty tabs.
    best_sheet, best_rows = None, []
    for ws in wb.worksheets:
        rows = []
        for r in ws.iter_rows(values_only=True):
            cells = ["" if v is None else str(v).strip() for v in r]
            if any(c for c in cells):  # skip fully blank rows
                rows.append(cells)
        if len(rows) > len(best_rows):
            best_sheet = ws.title
            best_rows = rows
    if not best_rows:
        raise ValidationError(
            "Workbook has no data rows.",
            code="COA004",
        )
    return best_rows, {
        "filename": filename,
        "format": "xlsx",
        "sheet": best_sheet,
        "row_count": len(best_rows),
    }


def _parse_csv(file_bytes: bytes, filename: str) -> tuple[list[list[Any]], dict]:
    text = file_bytes.decode("utf-8-sig", errors="replace")
    delim = "\t" if filename.lower().endswith(".tsv") else ","
    reader = csv.reader(io.StringIO(text), delimiter=delim)
    rows = [[(c or "").strip() for c in r] for r in reader if any(c for c in r)]
    if not rows:
        raise ValidationError("CSV has no data rows.", code="COA005")
    return rows, {
        "filename": filename,
        "format": "csv",
        "row_count": len(rows),
    }


def _ask_claude(rows: list[list[Any]]) -> list[dict]:
    """Send rows to Claude and parse the structured response."""
    try:
        import anthropic
    except ImportError as exc:
        raise ValidationError(
            "anthropic SDK is required for AI CoA import.",
            code="COA006",
        ) from exc
    api_key = settings.ANTHROPIC_API_KEY or None
    if not api_key:
        raise ValidationError(
            "ANTHROPIC_API_KEY is not configured.",
            code="COA007",
        )
    client = anthropic.Anthropic(api_key=api_key)
    model = getattr(settings, "CLAUDE_OCR_MODEL", "claude-sonnet-4-6")

    # Format rows as a fenced markdown table — keeps headers/columns
    # together and uses fewer tokens than JSON.
    table = "\n".join(["\t".join(cells) for cells in rows])
    system = _SYSTEM_PROMPT.format(
        types=", ".join(VALID_ACCOUNT_TYPES),
        subtypes=json.dumps(SUBTYPES_BY_TYPE, indent=2),
    )
    user_text = (
        "Below is a Chart of Accounts uploaded by the user. Each row is "
        "tab-separated cells from the source file. Headers may not be on "
        "the first row. Identify the columns, then call the extract_coa "
        "tool with one structured row per CoA entry.\n\n"
        f"```\n{table}\n```"
    )
    from .anthropic_throttle import claude_throttle, raise_friendly_rate_limit

    claude_throttle()
    try:
        message = client.messages.create(
            model=model,
            max_tokens=16000,
            system=system,
            tools=[_COA_TOOL],
            tool_choice={"type": "tool", "name": "extract_coa"},
            messages=[{"role": "user", "content": user_text}],
        )
    except anthropic.AuthenticationError as e:
        raise ValidationError(
            "ANTHROPIC_API_KEY is invalid or revoked.",
            code="COA008",
            details={"error": str(e)},
        )
    except anthropic.RateLimitError as e:
        raise_friendly_rate_limit(e, code="COA009_RATE_LIMIT")
    except anthropic.APIError as e:
        raise ValidationError(
            f"Claude API error: {e}",
            code="COA009",
            details={"error": str(e)},
        )
    tool_use = next(
        (b for b in message.content if getattr(b, "type", None) == "tool_use"),
        None,
    )
    if tool_use is None:
        raise ValidationError(
            f"Claude did not call extract_coa; stop_reason={message.stop_reason}.",
            code="COA010",
        )
    payload = dict(tool_use.input)
    return list(payload.get("accounts") or [])


_SYSTEM_PROMPT = """You are a precise bookkeeping assistant. The user uploaded
a Chart of Accounts file (Excel or CSV). Your job:

1. Locate the table — skip title rows / blank rows.
2. Identify which columns hold: code (gl number / account #), name, type
   (if present), parent (if present).
3. For every data row, infer the right `account_type` and `account_subtype`
   from the code prefix and name. Common conventions:
       - 1xxxx codes -> asset
       - 2xxxx codes -> liability
       - 3xxxx codes -> equity
       - 4xxxx codes -> revenue
       - 5xxxx-9xxxx codes -> expense
   Override the prefix when the name disambiguates (e.g. "Tax Refund Received"
   under 6xxxx is revenue, not expense).
4. Detect HEADER rows (e.g. "ASSETS", "Current Assets") that have no
   postings. Set is_header=true and leave subtype blank.
5. For each row, return a confidence score 0.0-1.0 and a one-sentence
   rationale. Be conservative — flag low confidence (< 0.7) when:
       - The code prefix conflicts with the name's intuition.
       - The type/subtype isn't an obvious fit to any of the allowed values.
       - You had to guess column meanings.

Allowed account_type values: {types}

Allowed account_subtype values per type (subtype MUST belong to the chosen
type's list):
{subtypes}

Return EVERY data row — do not collapse, group, or skip rows except blank/
title rows. Preserve the source code exactly.
"""
