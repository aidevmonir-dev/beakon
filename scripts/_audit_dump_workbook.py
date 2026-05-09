"""One-shot audit helper: dump headers, row counts, and full content for the
key blueprint sheets so we can compare against the codebase.

Run: python scripts/_audit_dump_workbook.py
Output goes to docs/blueprint/_audit_workbook_dump.md (gitignored if needed).
"""
from openpyxl import load_workbook
from pathlib import Path

WB_PATH = r"D:/Thomas/2026 04 17-DRAFT-CoA-Wealth management v2.xlsx"
OUT_PATH = Path(r"D:/bookkeeper/docs/blueprint/_audit_workbook_dump.md")

DATA_SHEETS = [
    "01 CoA Definition",
    "02 CoA Master",
    "03 CoA Mapping",
    "04 Dimensions Reference",
    "05 Dimension Values",
    "06 Controlled Lists",
    "07 Loan Master",
    "08 Instrument Master",
    "09 Dimension Validation Rules",
    "10_Counterparty_Master",
    "11_Related_Party_Master",
    "12_Bank_Account_Master",
    "13_Custodian_Master",
    "14_Portfolio_Master",
    "15_Property_Master",
    "16_Policy_Master",
    "17_Tax_Lot_Master",
]

FULL_DUMP_SHEETS = {
    "01 CoA Definition",
    "03 CoA Mapping",
    "04 Dimensions Reference",
    "06 Controlled Lists",
    "09 Dimension Validation Rules",
}

wb = load_workbook(WB_PATH, data_only=True, read_only=True)

lines = ["# Workbook audit dump", ""]
for name in DATA_SHEETS:
    if name not in wb.sheetnames:
        lines.append(f"## {name}\n\n_NOT FOUND_\n")
        continue
    ws = wb[name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        lines.append(f"## {name}\n\n_empty_\n")
        continue
    headers = rows[0]
    body = rows[1:]
    nonempty_body = [r for r in body if any(c not in (None, "") for c in r)]
    lines.append(f"## {name}")
    lines.append(f"Header row ({len([h for h in headers if h])} cols): {list(headers)}")
    lines.append(f"Non-empty data rows: {len(nonempty_body)}")
    if name in FULL_DUMP_SHEETS:
        lines.append("")
        lines.append("Full content:")
        for r in nonempty_body:
            lines.append(f"  - {list(r)}")
    else:
        lines.append("First 3 sample rows:")
        for r in nonempty_body[:3]:
            lines.append(f"  - {list(r)}")
    lines.append("")

OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
OUT_PATH.write_text("\n".join(lines), encoding="utf-8")
print(f"Wrote {OUT_PATH}")
